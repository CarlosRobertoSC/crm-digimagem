"""
CRM-Digimagem — Backend
Flask + SQLite. Sem dependências externas além do que já acompanha o Flask.

Como rodar:
    pip install flask --break-system-packages   # se ainda não tiver
    python3 app.py
Depois abra http://localhost:5000

Login de teste (criados no seed):
    admin1@digimagem.com   / admin123     (Admin Master 1)
    admin2@digimagem.com   / admin123     (Admin Master 2)
    carlos.vendas@lojadigimagem.com.br / vendas123   (Vendedor)
    ana@lojadigimagem.com.br           / vendas123   (Vendedora)
    tiago@lojadigimagem.com.br         / vendas123   (Vendedor)
"""

import sqlite3
import os
import uuid
import json
import re
import time
import threading
import base64
from collections import defaultdict, deque
import urllib.request
import urllib.error
from datetime import datetime, timedelta, timezone, date
from functools import wraps

from flask import Flask, request, jsonify, g, send_from_directory, Response
from werkzeug.security import generate_password_hash, check_password_hash
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired

# Caminho do banco: em produção, aponte para FORA da pasta do código
# (ex.: CRM_DB_PATH=/opt/crm/data/crm.db) — assim as atualizações do app
# nunca tocam nos dados.
DB_PATH = os.environ.get("CRM_DB_PATH", "crm.db")
SECRET_KEY = os.environ.get("SECRET_KEY", "troque-esta-chave-em-producao-para-algo-aleatorio")
if SECRET_KEY == "troque-esta-chave-em-producao-para-algo-aleatorio":
    print("[AVISO] Usando SECRET_KEY padrão de desenvolvimento. Defina a variável de "
          "ambiente SECRET_KEY com um valor aleatório antes de expor este app na internet.")
TOKEN_MAX_AGE = 60 * 60 * 24 * 7  # 7 dias

# O Assistente de Vendas é 100% GRATUITO: um motor de regras que roda localmente sobre
# os dados do seu próprio CRM (etapa do funil, tempo parado, tarefas pendentes, status
# do cliente). Não há nenhuma chamada a API externa nem a serviços pagos — não vai gerar
# custo nenhum, em nenhuma hipótese, agora ou depois.

serializer = URLSafeTimedSerializer(SECRET_KEY)
app = Flask(__name__, static_folder="static", static_url_path="")


# ------------------------------------------------------------------
# Infra de banco
# ------------------------------------------------------------------
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH, timeout=15)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
        g.db.execute("PRAGMA journal_mode = WAL")   # leituras não bloqueiam escritas
        g.db.execute("PRAGMA busy_timeout = 15000")
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def new_id():
    return uuid.uuid4().hex


def now_iso():
    return datetime.now(timezone.utc).replace(tzinfo=None).strftime("%Y-%m-%d %H:%M:%S")


def row_to_dict(row):
    return dict(row) if row else None


# ------------------------------------------------------------------
# Auth / RBAC
# ------------------------------------------------------------------
def make_token(user_id):
    return serializer.dumps({"uid": user_id})


def verify_token(token):
    try:
        data = serializer.loads(token, max_age=TOKEN_MAX_AGE)
        return data.get("uid")
    except (BadSignature, SignatureExpired):
        return None


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        auth = request.headers.get("Authorization", "")
        token = auth.replace("Bearer ", "").strip()
        uid = verify_token(token) if token else None
        if not uid:
            return jsonify({"error": "Não autenticado. Faça login novamente."}), 401
        db = get_db()
        user = db.execute(
            "SELECT * FROM users WHERE id = ? AND ativo = 1", (uid,)
        ).fetchone()
        if not user:
            return jsonify({"error": "Usuário não encontrado ou inativo."}), 401
        g.current_user = row_to_dict(user)
        return fn(*args, **kwargs)
    return wrapper


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if g.current_user["role"] != "admin":
            return jsonify({"error": "Apenas administradores podem executar esta ação."}), 403
        return fn(*args, **kwargs)
    return wrapper


def audit(acao, entidade, entidade_id=None, detalhes=None):
    db = get_db()
    db.execute(
        "INSERT INTO audit_log (id, user_id, acao, entidade, entidade_id, detalhes) VALUES (?,?,?,?,?,?)",
        (new_id(), g.current_user["id"], acao, entidade, entidade_id,
         json.dumps(detalhes) if detalhes else None),
    )


def _ultimo_dia_mes(ano, mes):
    if mes == 12:
        return date(ano, 12, 31)
    return date(ano, mes + 1, 1) - timedelta(days=1)


def resolve_period(tipo, offset=0, data_inicio=None, data_fim=None):
    """Resolve um período de datas a partir do tipo escolhido pelo usuário
    (quinzenal | mensal | trimestral | personalizado) e um deslocamento inteiro
    (offset) para navegar entre períodos anteriores/seguintes. Retorna (inicio, fim)
    como objetos date. Levanta ValueError se os parâmetros forem inválidos."""
    hoje = datetime.now(timezone.utc).replace(tzinfo=None).date()
    offset = int(offset or 0)

    if tipo == "personalizado":
        if not data_inicio or not data_fim:
            raise ValueError("Informe data_inicio e data_fim para o período personalizado.")
        inicio = datetime.strptime(data_inicio, "%Y-%m-%d").date()
        fim = datetime.strptime(data_fim, "%Y-%m-%d").date()
        if fim < inicio:
            raise ValueError("data_fim não pode ser anterior a data_inicio.")
        return inicio, fim

    if tipo == "quinzenal":
        quinzena_atual = 0 if hoje.day <= 15 else 1
        indice = (hoje.year * 24 + (hoje.month - 1) * 2 + quinzena_atual) + offset
        ano2 = indice // 24
        resto = indice % 24
        mes2 = resto // 2 + 1
        quinzena2 = resto % 2
        if quinzena2 == 0:
            return date(ano2, mes2, 1), date(ano2, mes2, 15)
        return date(ano2, mes2, 16), _ultimo_dia_mes(ano2, mes2)

    if tipo == "mensal":
        indice = (hoje.year * 12 + (hoje.month - 1)) + offset
        ano2 = indice // 12
        mes2 = indice % 12 + 1
        return date(ano2, mes2, 1), _ultimo_dia_mes(ano2, mes2)

    if tipo == "trimestral":
        trimestre_atual = (hoje.month - 1) // 3
        indice = (hoje.year * 4 + trimestre_atual) + offset
        ano2 = indice // 4
        trimestre2 = indice % 4
        mes_inicio = trimestre2 * 3 + 1
        mes_fim = mes_inicio + 2
        return date(ano2, mes_inicio, 1), _ultimo_dia_mes(ano2, mes_fim)

    raise ValueError("tipo de período inválido. Use: quinzenal, mensal, trimestral ou personalizado.")


def parse_period_from_request():
    """Lê os parâmetros de período da query string e devolve (inicio_str, fim_str)
    prontos para usar em cláusulas SQL BETWEEN — ou (None, None) se nenhum período
    foi pedido (relatório mostra a situação atual, sem filtro de data)."""
    tipo = request.args.get("periodo")
    if not tipo:
        return None, None
    offset = request.args.get("offset", 0)
    data_inicio = request.args.get("data_inicio")
    data_fim = request.args.get("data_fim")
    inicio, fim = resolve_period(tipo, offset, data_inicio, data_fim)
    return inicio.strftime("%Y-%m-%d"), fim.strftime("%Y-%m-%d")


def report_scope_clause(field="user_id"):
    """Regra de visibilidade dos relatórios: vendedor só vê o próprio; admin vê o
    próprio por padrão, um vendedor específico via ?vendedor_id=, ou todo mundo via
    ?scope=all. Retorna (clause, params)."""
    if g.current_user["role"] != "admin":
        return f" AND {field} = ?", [g.current_user["id"]]
    if request.args.get("scope") == "all":
        return "", []
    vendedor_id = request.args.get("vendedor_id")
    if vendedor_id:
        return f" AND {field} = ?", [vendedor_id]
    return f" AND {field} = ?", [g.current_user["id"]]


def scope_filter_clause(field="responsavel_id"):
    """Retorna (clause, params) para restringir vendedor aos próprios registros.
    Admin só vê tudo se pedir explicitamente ?scope=all."""
    if g.current_user["role"] == "admin" and request.args.get("scope") == "all":
        return "", []
    if g.current_user["role"] == "admin":
        # admin sem scope=all → vê os dados como se fosse ele mesmo (padrão "meus dados")
        return f" AND {field} = ?", [g.current_user["id"]]
    return f" AND {field} = ?", [g.current_user["id"]]


def active_admin_count(db, exclude_user_id=None):
    q = "SELECT COUNT(*) c FROM users WHERE role='admin' AND ativo=1"
    params = []
    if exclude_user_id:
        q += " AND id != ?"
        params.append(exclude_user_id)
    return db.execute(q, params).fetchone()["c"]


# ------------------------------------------------------------------
# Motor de Alertas (regras sobre dados estruturados — sem leitura de conversas)
# ------------------------------------------------------------------
# ------------------------------------------------------------------
# CPF/CNPJ — com suporte ao CNPJ ALFANUMÉRICO (IN RFB nº 2.229/2024,
# vigente desde julho/2026): 12 posições [0-9A-Z] + 2 dígitos verificadores
# numéricos. O DV usa módulo 11 com valor do caractere = ASCII - 48
# (dígitos mantêm seu valor; A=17 … Z=42), o que torna o MESMO cálculo
# válido para os CNPJs numéricos antigos (retrocompatível).
# Vetor de teste oficial da Receita: 12ABC34501DE -> DV 35.
# ------------------------------------------------------------------
CNPJ_RE = re.compile(r"^[0-9A-Z]{12}[0-9]{2}$")


def normalizar_cpf_cnpj(valor):
    """Remove a máscara (pontos, traço, barra, espaços) e põe em maiúsculas.
    NUNCA remover letras: em CNPJs novos elas fazem parte do número."""
    return re.sub(r"[.\-/\s]", "", str(valor or "")).upper()


def _dv_cnpj(base):
    """Um dígito verificador do CNPJ (base com 12 ou 13 caracteres):
    pesos 2..9 aplicados da direita para a esquerda, em ciclo."""
    soma = 0
    for i, ch in enumerate(reversed(base)):
        soma += (ord(ch) - 48) * (2 + (i % 8))
    resto = soma % 11
    return 0 if resto < 2 else 11 - resto


def cnpj_valido(cnpj):
    if not CNPJ_RE.match(cnpj):
        return False
    dv1 = _dv_cnpj(cnpj[:12])
    dv2 = _dv_cnpj(cnpj[:12] + str(dv1))
    return cnpj[12:] == f"{dv1}{dv2}"


def cpf_valido(cpf):
    if not re.fullmatch(r"[0-9]{11}", cpf) or cpf == cpf[0] * 11:
        return False
    for n in (9, 10):
        soma = sum(int(cpf[i]) * ((n + 1) - i) for i in range(n))
        if (soma * 10) % 11 % 10 != int(cpf[n]):
            return False
    return True


def validar_cpf_cnpj_ou_erro(valor):
    """Normaliza e valida. Retorna (valor_normalizado, None) quando ok,
    (None, None) quando vazio (campo é opcional), ou (None, mensagem)."""
    v = normalizar_cpf_cnpj(valor)
    if not v:
        return None, None
    if len(v) == 11 and v.isdigit():
        return (v, None) if cpf_valido(v) else (None, "CPF inválido — confira os dígitos.")
    if len(v) == 14:
        if cnpj_valido(v):
            return v, None
        return None, "CNPJ inválido — o dígito verificador não confere (lembre: CNPJs novos podem conter letras maiúsculas)."
    return None, "CPF/CNPJ inválido — use 11 caracteres (CPF) ou 14 (CNPJ)."


# ------------------------------------------------------------------
# 🖨 Equipamentos Fujifilm que o cliente possui (campo estruturado).
# Os marcados como compatíveis alimentam a "cobertura de oferta" do
# software Revele Momentos.
# ------------------------------------------------------------------
EQUIPAMENTOS = {
    "ask300": "ASK-300", "ask400": "ASK-400",
    "dx100": "Frontier DX100", "de100": "Frontier DE100",
    "de100xd": "Frontier DE100-XD", "dx400": "Frontier Smartlab DX400",
    "minilab": "Minilab (química)",
    "instax": "Instax (câmeras e filmes)",
}
EQUIPAMENTOS_COMPATIVEIS_SOFTWARE = ("ask300", "ask400", "dx100", "de100", "de100xd", "dx400")


def normalizar_equipamentos(valor):
    """Aceita lista ou JSON string; mantém só slugs conhecidos.
    Retorna JSON string ou None (nenhum equipamento)."""
    itens = valor
    if isinstance(valor, str):
        try:
            itens = json.loads(valor)
        except Exception:
            itens = []
    if not isinstance(itens, list):
        itens = []
    itens = [i for i in itens if i in EQUIPAMENTOS]
    return json.dumps(itens) if itens else None


def _clientes_sem_oferta_sql(alias="c"):
    """(fragmento_like, params) para clientes com equipamento compatível e
    nenhum negócio de software jamais criado."""
    likes = " OR ".join(f"{alias}.equipamentos LIKE ?" for _ in EQUIPAMENTOS_COMPATIVEIS_SOFTWARE)
    params = [f'%"{s}"%' for s in EQUIPAMENTOS_COMPATIVEIS_SOFTWARE]
    return likes, params


def normalizar_telefone(v):
    """Telefones são guardados apenas com dígitos (DDI+DDD+número) —
    exatamente o formato que o wa.me e futuras integrações exigem."""
    d = re.sub(r"\D", "", str(v or ""))
    return d or None


def _int_or_none(v):
    """Converte para inteiro positivo, ou None (usado no ciclo de recompra)."""
    try:
        n = int(str(v).strip())
        return n if n > 0 else None
    except (TypeError, ValueError):
        return None


def _hoje_mais_dias(dias):
    return (datetime.now(timezone.utc) + timedelta(days=dias)).strftime("%Y-%m-%d")


def run_alert_engine():
    db = get_db()

    # 1) Inatividade — clientes recorrentes/vip sem comprar há mais de 30 dias
    limite_inatividade = (datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=30)).strftime("%Y-%m-%d")
    inativos = db.execute("""
        SELECT * FROM customers
        WHERE status_fidelidade IN ('recorrente','vip')
          AND data_ultima_compra IS NOT NULL
          AND data_ultima_compra < ?
          AND ativo = 1
    """, (limite_inatividade,)).fetchall()
    for c in inativos:
        dias = (datetime.now(timezone.utc).replace(tzinfo=None) - datetime.strptime(c["data_ultima_compra"], "%Y-%m-%d")).days
        _upsert_insight(
            customer_id=c["id"], deal_id=None, user_id=c["responsavel_id"],
            tipo="inatividade", prioridade="urgente" if dias > 45 else "alta",
            descricao=f'{c["nome"]} está há {dias} dias sem comprar (cliente {c["status_fidelidade"]}).',
            dados_extra={"dias_sem_comprar": dias},
        )

    # 2) Abandono de funil — deals abertos parados há mais de 48h na mesma etapa
    limite_funil = (datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(hours=48)).strftime("%Y-%m-%d %H:%M:%S")
    parados = db.execute("""
        SELECT d.*, c.nome as cliente_nome FROM deals d
        JOIN customers c ON c.id = d.customer_id
        WHERE d.status = 'aberto' AND d.etapa_atualizada_em < ? AND c.ativo = 1
    """, (limite_funil,)).fetchall()
    for d in parados:
        horas = int((datetime.now(timezone.utc).replace(tzinfo=None) - datetime.strptime(d["etapa_atualizada_em"], "%Y-%m-%d %H:%M:%S")).total_seconds() // 3600)
        _upsert_insight(
            customer_id=d["customer_id"], deal_id=d["id"], user_id=d["user_id"],
            tipo="abandono_funil", prioridade="urgente" if horas > 72 else "alta",
            descricao=f'Negócio "{d["titulo"]}" com {d["cliente_nome"]} parado em "{d["etapa_funil"]}" há {horas}h.',
            dados_extra={"horas_parado": horas, "etapa": d["etapa_funil"]},
        )

    # 3) Compromisso esquecido — tasks vencidas e não concluídas
    agora = now_iso()
    vencidas = db.execute("""
        SELECT t.*, c.nome as cliente_nome FROM tasks t
        LEFT JOIN customers c ON c.id = t.customer_id
        WHERE t.executado = 0 AND t.data_lembrete < ? AND (c.ativo = 1 OR c.ativo IS NULL)
    """, (agora,)).fetchall()
    for t in vencidas:
        nome_cliente = t["cliente_nome"] or "cliente"
        _upsert_insight(
            customer_id=t["customer_id"], deal_id=t["deal_id"], user_id=t["user_id"],
            tipo="compromisso_esquecido", prioridade="urgente",
            descricao=f'Compromisso não cumprido com {nome_cliente}: "{t["descricao"]}".',
            dados_extra={"task_id": t["id"]},
        )
        # escalonamento: se o compromisso venceu há mais de 24h, marca para admins verem também
        vencido_em = datetime.strptime(t["data_lembrete"], "%Y-%m-%d %H:%M:%S")
        if (datetime.now(timezone.utc).replace(tzinfo=None) - vencido_em) > timedelta(hours=24):
            db.execute(
                "UPDATE ai_insights SET escalado_admin = 1 WHERE customer_id = ? AND tipo_alerta = 'compromisso_esquecido' AND lido = 0",
                (t["customer_id"],),
            )

    # 4) Sugestão de conteúdo — clientes VIP sem alerta de fidelização recente
    vips = db.execute("SELECT * FROM customers WHERE status_fidelidade = 'vip' AND ativo = 1").fetchall()
    for c in vips:
        existe = db.execute("""
            SELECT 1 FROM ai_insights WHERE customer_id = ? AND tipo_alerta = 'sugestao_conteudo'
        """, (c["id"],)).fetchone()
        if not existe:
            _upsert_insight(
                customer_id=c["id"], deal_id=None, user_id=c["responsavel_id"],
                tipo="sugestao_conteudo", prioridade="media",
                descricao=f'{c["nome"]} é cliente VIP — considere enviar catálogo atualizado ou um mimo de fidelização.',
                dados_extra=None,
            )

    # 5) 🔁 Recompra programada — quando a data agendada do cliente chega, cria
    # automaticamente o novo negócio (valor ZERO, para o vendedor preencher ao
    # negociar) e um compromisso que aparece no 📌 do Feed de Alertas do
    # responsável. Antiduplicação: só gera se o cliente NÃO tiver nenhum
    # negócio aberto — o ciclo continua quando o negócio atual for fechado.
    # Gera até 1 dia ANTES da data de recompra, para o lembrete aparecer no
    # feed no máximo 24 horas antes do contato (regra do negócio).
    ate_amanha_recompra = (datetime.now(timezone.utc) + timedelta(days=1)).strftime("%Y-%m-%d")
    devidos = db.execute("""
        SELECT * FROM customers
        WHERE recompra_dias IS NOT NULL AND recompra_dias > 0
          AND proxima_recompra IS NOT NULL AND proxima_recompra <= ?
          AND ativo = 1 AND responsavel_id IS NOT NULL
          AND NOT EXISTS (
              SELECT 1 FROM deals d WHERE d.customer_id = customers.id AND d.status = 'aberto'
          )
    """, (ate_amanha_recompra,)).fetchall()
    for c in devidos:
        did = new_id()
        db.execute("""
            INSERT INTO deals (id, customer_id, user_id, titulo, etapa_funil,
                               valor_estimado, status, origem_recompra)
            VALUES (?,?,?,?,'novo_lead',0,'aberto',1)
        """, (did, c["id"], c["responsavel_id"], "Recompra de insumos"))
        db.execute("""
            INSERT INTO deal_stage_history (id, deal_id, etapa_anterior, etapa_nova, user_id, data_transicao)
            VALUES (?,?,NULL,'novo_lead',?,?)
        """, (new_id(), did, c["responsavel_id"], now_iso()))
        db.execute("""
            INSERT INTO deal_value_history (id, deal_id, valor_anterior, valor_novo, user_id, created_at)
            VALUES (?,?,NULL,0,?,?)
        """, (new_id(), did, c["responsavel_id"], now_iso()))
        db.execute("""
            INSERT INTO tasks (id, customer_id, deal_id, user_id, descricao, tipo_atividade, data_lembrete)
            VALUES (?,?,?,?,?,'follow_up',?)
        """, (new_id(), c["id"], did, c["responsavel_id"],
              f'🔁 Recompra programada: contatar {c["nome"]} para nova venda de insumos',
              f'{c["proxima_recompra"]} 12:00:00'))
        # desarma a agenda; o ciclo é reprogramado quando este negócio for fechado
        db.execute("UPDATE customers SET proxima_recompra = NULL WHERE id = ?", (c["id"],))
        audit("create", "deals", did, {"origem": "recompra_programada", "customer_id": c["id"]})

    db.commit()


def _upsert_insight(customer_id, deal_id, user_id, tipo, prioridade, descricao, dados_extra):
    db = get_db()
    # Não duplica se já há alerta ATIVO do mesmo tipo para o cliente, nem se um
    # alerta desse tipo foi criado nos últimos 7 dias (mesmo que marcado como
    # lido) — "Marcar lido" vale como um silêncio de 7 dias enquanto a condição
    # persistir; se ela continuar depois disso, o alerta volta.
    limite_silencio = (datetime.now(timezone.utc) - timedelta(days=7)).strftime("%Y-%m-%d %H:%M:%S")
    existente = db.execute("""
        SELECT id FROM ai_insights
        WHERE customer_id = ? AND tipo_alerta = ?
          AND (lido = 0 OR created_at >= ?)
    """, (customer_id, tipo, limite_silencio)).fetchone()
    if existente:
        return
    db.execute("""
        INSERT INTO ai_insights (id, customer_id, deal_id, user_id, tipo_alerta, prioridade, descricao_insight, dados_extra)
        VALUES (?,?,?,?,?,?,?,?)
    """, (new_id(), customer_id, deal_id, user_id, tipo, prioridade, descricao,
          json.dumps(dados_extra) if dados_extra else None))


# ------------------------------------------------------------------
# Assistente de Vendas — motor de regras 100% local e gratuito.
# Não chama nenhuma API externa nem serviço pago: só lê os dados que já
# estão no seu banco (etapa do funil, tempo parado, tarefas, cliente) e
# aplica um checklist de boas práticas de vendas.
# ------------------------------------------------------------------
ETAPA_PESO = {
    "novo_lead": 15, "qualificacao": 35, "proposta_enviada": 55,
    "negociacao": 75, "fechado_ganho": 100, "fechado_perdido": 0,
}

DICAS_GERAIS = [
    "Responda leads novos em até algumas horas — a chance de conversão despenca depois do primeiro dia.",
    "Sempre marque um próximo passo com data antes de encerrar uma conversa com o cliente.",
    "Clientes VIP fecham mais rápido quando recebem algum diferencial exclusivo (condição, brinde, prioridade de agenda).",
    "Negócios parados há mais de 48h numa etapa quase sempre precisam de um empurrão ativo, não de esperar o cliente responder.",
    "Depois de enviar uma proposta, confirme em 24h se o cliente teve alguma dúvida — silêncio geralmente é objeção não dita.",
]


def _hours_since(iso_ts):
    dt = datetime.strptime(iso_ts, "%Y-%m-%d %H:%M:%S")
    return (datetime.now(timezone.utc).replace(tzinfo=None) - dt).total_seconds() / 3600


def build_deal_recommendation(deal, customer, tasks):
    """Retorna (score 0-100, lista de recomendações) para um negócio, sem nenhuma
    chamada externa — só regras sobre os dados estruturados do próprio negócio."""
    horas_parado = _hours_since(deal["etapa_atualizada_em"])
    horas_criado = _hours_since(deal["created_at"])
    etapa = deal["etapa_funil"]
    score = ETAPA_PESO.get(etapa, 20)

    tarefas_abertas = [t for t in tasks if not t["executado"]]
    tarefas_vencidas = [t for t in tarefas_abertas
                         if datetime.strptime(t["data_lembrete"], "%Y-%m-%d %H:%M:%S") < datetime.now(timezone.utc).replace(tzinfo=None)]

    recomendacoes = []

    if horas_parado > 96:
        score -= 30
        recomendacoes.append({"prioridade": "urgente",
            "texto": f"Parado há {int(horas_parado)}h nesta etapa — está bem acima do razoável. Retome contato hoje ou reavalie se este negócio ainda é viável."})
    elif horas_parado > 48:
        score -= 15
        recomendacoes.append({"prioridade": "alta",
            "texto": f"Parado há {int(horas_parado)}h na etapa atual (limite saudável: 48h). Faça contato ativo em vez de esperar o cliente responder."})

    if tarefas_vencidas:
        score -= 20
        recomendacoes.append({"prioridade": "urgente",
            "texto": f'Existe {len(tarefas_vencidas)} compromisso(s) vencido(s) com este cliente. Resolva isso antes de qualquer outra ação — é o que mais derruba confiança.'})
    elif not tarefas_abertas:
        score -= 10
        recomendacoes.append({"prioridade": "alta",
            "texto": "Não há nenhum próximo passo agendado para este negócio. Marque um follow-up agora — negócio sem próxima data tende a esfriar."})

    if etapa == "novo_lead" and horas_criado > 24:
        score -= 10
        recomendacoes.append({"prioridade": "alta",
            "texto": "Lead novo há mais de 24h sem avançar para qualificação. Priorize o primeiro contato — a taxa de resposta cai muito depois do primeiro dia."})

    if etapa == "proposta_enviada" and horas_parado > 24:
        recomendacoes.append({"prioridade": "media",
            "texto": "Proposta enviada há mais de 24h sem retorno. Confirme se o cliente teve alguma dúvida — silêncio costuma ser objeção de preço não dita."})

    if etapa == "negociacao":
        recomendacoes.append({"prioridade": "media",
            "texto": "Você está na reta final. Reforce os diferenciais do serviço e, se fizer sentido, ofereça uma condição para fechar ainda esta semana."})

    if customer["status_fidelidade"] == "vip":
        score += 10
        recomendacoes.append({"prioridade": "baixa",
            "texto": "Cliente VIP — considere um diferencial exclusivo (condição especial, prioridade de agenda ou um mimo) para acelerar a decisão."})

    if deal["valor_estimado"] and deal["valor_estimado"] >= 10000:
        recomendacoes.append({"prioridade": "baixa",
            "texto": "Negócio de ticket alto — vale a pena uma conversa por ligação ou vídeo em vez de só texto, para reduzir o risco de mal-entendido."})

    if not recomendacoes:
        recomendacoes.append({"prioridade": "baixa",
            "texto": "Este negócio está em dia: sem pendências vencidas e dentro do tempo esperado na etapa. Continue o ritmo combinado com o cliente."})

    score = max(0, min(100, round(score)))
    ordem_prioridade = {"urgente": 0, "alta": 1, "media": 2, "baixa": 3}
    recomendacoes.sort(key=lambda r: ordem_prioridade[r["prioridade"]])
    return score, recomendacoes


# ------------------------------------------------------------------
# Rate limiting do login — proteção contra força bruta.
# Em memória, o que é adequado ao deploy atual (1 worker no gunicorn).
# Se um dia escalar para múltiplos workers/instâncias, mover o estado
# para Redis ou equivalente.
# ------------------------------------------------------------------
LOGIN_MAX_FALHAS = 5            # tentativas erradas permitidas…
LOGIN_JANELA_SEGUNDOS = 15 * 60  # …dentro desta janela (15 min)
_login_falhas = defaultdict(deque)   # ip -> timestamps (monotonic) das falhas
_login_lock = threading.Lock()


def _client_ip():
    """IP real do cliente. Atrás do proxy do Render/Heroku ele vem no
    X-Forwarded-For (primeiro endereço da lista)."""
    xff = request.headers.get("X-Forwarded-For", "")
    if xff:
        return xff.split(",")[0].strip()
    return request.remote_addr or "desconhecido"


def _login_bloqueado(ip):
    """Retorna quantos segundos faltam para o IP poder tentar de novo,
    ou 0 se ainda está dentro do limite."""
    agora = time.monotonic()
    with _login_lock:
        falhas = _login_falhas[ip]
        while falhas and agora - falhas[0] > LOGIN_JANELA_SEGUNDOS:
            falhas.popleft()
        if len(falhas) >= LOGIN_MAX_FALHAS:
            return int(LOGIN_JANELA_SEGUNDOS - (agora - falhas[0])) + 1
    return 0


def _login_registrar_falha(ip):
    with _login_lock:
        _login_falhas[ip].append(time.monotonic())


def _login_limpar(ip):
    with _login_lock:
        _login_falhas.pop(ip, None)


# ------------------------------------------------------------------
# Rotas — Auth
# ------------------------------------------------------------------
@app.post("/api/auth/login")
def login():
    ip = _client_ip()
    espera = _login_bloqueado(ip)
    if espera:
        minutos = max(1, -(-espera // 60))  # arredonda pra cima
        resp = jsonify({"error": f"Muitas tentativas de login. Aguarde {minutos} minuto(s) e tente novamente."})
        resp.headers["Retry-After"] = str(espera)
        return resp, 429
    body = request.get_json(force=True, silent=True) or {}
    email = (body.get("email") or "").strip().lower()
    senha = body.get("senha") or ""
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE email = ? AND ativo = 1", (email,)).fetchone()
    if not user or not check_password_hash(user["senha_hash"], senha):
        _login_registrar_falha(ip)
        return jsonify({"error": "Email ou senha inválidos."}), 401
    _login_limpar(ip)
    token = make_token(user["id"])
    return jsonify({
        "token": token,
        "user": {"id": user["id"], "nome": user["nome"], "role": user["role"], "email": user["email"]},
    })


@app.get("/api/me")
@login_required
def me():
    u = g.current_user
    return jsonify({"id": u["id"], "nome": u["nome"], "role": u["role"], "email": u["email"]})


# ------------------------------------------------------------------
# Rotas — Dashboard (stats agregadas)
# ------------------------------------------------------------------
@app.get("/api/dashboard/stats")
@login_required
def dashboard_stats():
    run_alert_engine()
    db = get_db()
    clause, params = scope_filter_clause("user_id")

    # alertas urgentes visíveis no escopo do usuário
    if g.current_user["role"] == "admin" and request.args.get("scope") == "all":
        urgentes = db.execute(
            "SELECT COUNT(*) c FROM ai_insights WHERE lido = 0 AND prioridade = 'urgente'"
        ).fetchone()["c"]
        valor_aberto = db.execute(
            "SELECT COALESCE(SUM(valor_estimado),0) v FROM deals WHERE status='aberto'"
        ).fetchone()["v"]
        parados = db.execute(
            "SELECT COUNT(*) c FROM deals WHERE status='aberto' AND etapa_atualizada_em < ?",
            ((datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(hours=48)).strftime("%Y-%m-%d %H:%M:%S"),)
        ).fetchone()["c"]
    else:
        uid = g.current_user["id"]
        urgentes = db.execute(
            "SELECT COUNT(*) c FROM ai_insights WHERE lido = 0 AND prioridade = 'urgente' AND user_id = ?", (uid,)
        ).fetchone()["c"]
        valor_aberto = db.execute(
            "SELECT COALESCE(SUM(valor_estimado),0) v FROM deals WHERE status='aberto' AND user_id = ?", (uid,)
        ).fetchone()["v"]
        parados = db.execute(
            "SELECT COUNT(*) c FROM deals WHERE status='aberto' AND user_id = ? AND etapa_atualizada_em < ?",
            (uid, (datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(hours=48)).strftime("%Y-%m-%d %H:%M:%S"))
        ).fetchone()["c"]

    return jsonify({"alertas_urgentes": urgentes, "valor_em_negociacao": valor_aberto, "deals_parados": parados})


# ------------------------------------------------------------------
# Rotas — Alertas (ai_insights)
# ------------------------------------------------------------------
@app.get("/api/insights")
@login_required
def list_insights():
    run_alert_engine()
    db = get_db()
    if g.current_user["role"] == "admin" and request.args.get("scope") == "all":
        rows = db.execute("""
            SELECT i.*, c.nome as cliente_nome, u.nome as vendedor_nome FROM ai_insights i
            JOIN customers c ON c.id = i.customer_id
            LEFT JOIN users u ON u.id = i.user_id
            WHERE i.lido = 0
            ORDER BY CASE i.prioridade WHEN 'urgente' THEN 0 WHEN 'alta' THEN 1 WHEN 'media' THEN 2 ELSE 3 END, i.created_at DESC
        """).fetchall()
    else:
        # Cada usuário só vê os próprios alertas — mesmo os escalados (compromisso vencido
        # há +24h) não vazam para outros vendedores; o admin já enxerga tudo via scope=all.
        rows = db.execute("""
            SELECT i.*, c.nome as cliente_nome FROM ai_insights i
            JOIN customers c ON c.id = i.customer_id
            WHERE i.lido = 0 AND i.user_id = ?
            ORDER BY CASE i.prioridade WHEN 'urgente' THEN 0 WHEN 'alta' THEN 1 WHEN 'media' THEN 2 ELSE 3 END, i.created_at DESC
        """, (g.current_user["id"],)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/insights/<insight_id>/read")
@login_required
def mark_insight_read(insight_id):
    db = get_db()
    existing = db.execute("SELECT * FROM ai_insights WHERE id = ?", (insight_id,)).fetchone()
    if not existing:
        return jsonify({"error": "Alerta não encontrado."}), 404
    if g.current_user["role"] != "admin" and existing["user_id"] != g.current_user["id"]:
        return jsonify({"error": "Sem permissão para alterar este alerta."}), 403
    db.execute("UPDATE ai_insights SET lido = 1 WHERE id = ?", (insight_id,))
    audit("update", "ai_insights", insight_id, {"lido": True})
    db.commit()
    return jsonify({"ok": True})


# ------------------------------------------------------------------
# Rotas — Customers
# ------------------------------------------------------------------
@app.get("/api/customers")
@login_required
def list_customers():
    db = get_db()
    clause, params = scope_filter_clause("responsavel_id")
    params = list(params)
    filtro_ativo = "" if request.args.get("incluir_inativos") == "1" else " AND c.ativo = 1"

    # Busca livre (?q=) por nome, email, CPF/CNPJ, cidade, WhatsApp ou telefone.
    # Curingas do LIKE (% e _) são escapados para a busca ser sempre literal.
    filtro_busca = ""
    q = (request.args.get("q") or "").strip()
    if q:
        q_like = "%" + q.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_") + "%"
        filtro_busca = (
            " AND (c.nome LIKE ? ESCAPE '\\' OR c.email LIKE ? ESCAPE '\\'"
            " OR c.cpf_cnpj LIKE ? ESCAPE '\\' OR c.cidade LIKE ? ESCAPE '\\'"
            " OR c.whatsapp_id LIKE ? ESCAPE '\\' OR c.telefone LIKE ? ESCAPE '\\')"
        )
        params += [q_like] * 6

    # Filtro por equipamento (?equipamento=ask400 etc.)
    filtro_equip = ""
    eq = (request.args.get("equipamento") or "").strip()
    if eq in EQUIPAMENTOS:
        filtro_equip = " AND c.equipamentos LIKE ?"
        params.append(f'%"{eq}"%')

    # Paginação opcional (?limit=&offset=). Sem limit, mantém o comportamento
    # atual de retornar tudo — os seletores de cliente do frontend dependem disso.
    try:
        limit = int(request.args.get("limit", 0))
        offset = int(request.args.get("offset", 0))
    except (TypeError, ValueError):
        limit, offset = 0, 0
    clausula_limite = ""
    if limit > 0:
        clausula_limite = " LIMIT ? OFFSET ?"
        params += [limit, max(0, offset)]

    rows = db.execute(f"""
        SELECT c.*, u.nome as responsavel_nome FROM customers c
        LEFT JOIN users u ON u.id = c.responsavel_id
        WHERE 1=1 {clause}{filtro_ativo}{filtro_busca}{filtro_equip}
        ORDER BY c.nome{clausula_limite}
    """, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.get("/api/customers/<customer_id>")
@login_required
def get_customer(customer_id):
    db = get_db()
    c = db.execute("""
        SELECT c.*, u.nome as responsavel_nome FROM customers c
        LEFT JOIN users u ON u.id = c.responsavel_id
        WHERE c.id = ?
    """, (customer_id,)).fetchone()
    if not c:
        return jsonify({"error": "Cliente não encontrado."}), 404
    if g.current_user["role"] != "admin" and c["responsavel_id"] != g.current_user["id"]:
        return jsonify({"error": "Sem permissão para ver este cliente."}), 403
    # ⇄ transferências: admin vê de quem/para quem; os demais só data e admin
    eh_admin = g.current_user["role"] == "admin"
    transfers = [{"data": t["created_at"], "admin_nome": t["admin_nome"],
                  "de_nome": t["de_nome"] if eh_admin else None,
                  "para_nome": t["para_nome"] if eh_admin else None}
                 for t in db.execute("""
                     SELECT t.created_at, du.nome as de_nome, pu.nome as para_nome, au.nome as admin_nome
                     FROM customer_transfers t
                     LEFT JOIN users du ON du.id = t.de_user_id
                     LEFT JOIN users pu ON pu.id = t.para_user_id
                     LEFT JOIN users au ON au.id = t.admin_id
                     WHERE t.customer_id = ? ORDER BY t.created_at DESC
                 """, (customer_id,)).fetchall()]
    whatsapp_compartilhado = []
    if eh_admin and c["whatsapp_id"]:
        whatsapp_compartilhado = [{"nome": x["nome"], "responsavel_nome": x["responsavel_nome"]}
                                  for x in db.execute("""
                                      SELECT c2.nome, u2.nome as responsavel_nome
                                      FROM customers c2 LEFT JOIN users u2 ON u2.id = c2.responsavel_id
                                      WHERE c2.whatsapp_id = ? AND c2.id != ?
                                  """, (c["whatsapp_id"], customer_id)).fetchall()]
    deals = db.execute("SELECT * FROM deals WHERE customer_id = ? ORDER BY created_at DESC", (customer_id,)).fetchall()
    tasks = db.execute("SELECT * FROM tasks WHERE customer_id = ? ORDER BY data_lembrete", (customer_id,)).fetchall()
    insights = db.execute("SELECT * FROM ai_insights WHERE customer_id = ? AND lido = 0 ORDER BY created_at DESC", (customer_id,)).fetchall()
    audit("read", "customers", customer_id)
    db.commit()
    return jsonify({
        "customer": dict(c),
        "deals": [dict(d) for d in deals],
        "tasks": [dict(t) for t in tasks],
        "insights": [dict(i) for i in insights],
        "transfers": transfers,
        "whatsapp_compartilhado": whatsapp_compartilhado,
    })


def _erro_cliente_duplicado(db, cpf_cnpj, whatsapp, ignorar_id=None):
    """Trava de duplicidade da carteira, com semânticas diferentes:
    - CPF/CNPJ identifica a EMPRESA: repetido é duplicata real — bloqueia
      sempre (na própria carteira ou na alheia).
    - WhatsApp identifica a PESSOA, que pode ter várias empresas: repetido
      na PRÓPRIA carteira é permitido (2ª empresa do mesmo cliente);
      bloqueia apenas se o número existir EXCLUSIVAMENTE em carteira de
      outro usuário — aí o caminho é falar com o administrador."""
    uid = g.current_user["id"]

    def _bloqueio(achado):
        if achado["responsavel_id"] == uid:
            return jsonify({"error": f"Você já tem este cliente na sua carteira: \"{achado['nome']}\"."}), 409
        if g.current_user["role"] == "admin":
            dono = db.execute("SELECT nome FROM users WHERE id = ?", (achado["responsavel_id"],)).fetchone()
            return jsonify({"error": f"Este cliente já está cadastrado no sistema: \"{achado['nome']}\", na carteira de {dono['nome'] if dono else '?'}. Se for o caso, abra a ficha dele e use ⇄ Transferir."}), 409
        return jsonify({"error": "Este cliente já foi cadastrado no sistema por outro usuário. Fale com o administrador sobre este cliente."}), 409

    if cpf_cnpj:
        achado = db.execute("SELECT * FROM customers WHERE cpf_cnpj = ? AND id != ?",
                            (cpf_cnpj, ignorar_id or "")).fetchone()
        if achado:
            return _bloqueio(achado)
    if whatsapp:
        iguais = db.execute("SELECT * FROM customers WHERE whatsapp_id = ? AND id != ?",
                            (whatsapp, ignorar_id or "")).fetchall()
        if iguais and not any(x["responsavel_id"] == uid for x in iguais):
            # o número só existe em carteira alheia -> bloqueia
            return _bloqueio(iguais[0])
    return None


def _ancora_recompra(data_base, dias):
    """🔁 Próximo contato = ÚLTIMA COMPRA + ciclo (âncora comercial correta).
    Sem última compra registrada, a âncora é hoje. Se a conta cair no passado
    (cliente já atrasado no ciclo), a data é mantida: o motor de recompra
    gera o contato imediatamente — que é o comportamento certo."""
    if not dias:
        return None
    base = None
    if data_base:
        try:
            base = datetime.strptime(str(data_base)[:10], "%Y-%m-%d")
        except ValueError:
            base = None
    if base is None:
        return _hoje_mais_dias(dias)
    return (base + timedelta(days=dias)).strftime("%Y-%m-%d")


@app.post("/api/customers")
@login_required
def create_customer():
    body = request.get_json(force=True, silent=True) or {}
    if not body.get("nome"):
        return jsonify({"error": "Nome é obrigatório."}), 400
    db = get_db()
    cid = new_id()
    # Vendedor só pode cadastrar cliente para si mesmo; só admin pode atribuir a outro usuário.
    responsavel_id = body.get("responsavel_id", g.current_user["id"])
    if g.current_user["role"] != "admin":
        responsavel_id = g.current_user["id"]
    cpf_cnpj_norm, erro_doc = validar_cpf_cnpj_ou_erro(body.get("cpf_cnpj"))
    if erro_doc:
        return jsonify({"error": erro_doc}), 400
    erro_dup = _erro_cliente_duplicado(db, cpf_cnpj_norm, normalizar_telefone(body.get("whatsapp_id")))
    if erro_dup:
        return erro_dup
    recompra_dias = _int_or_none(body.get("recompra_dias"))
    db.execute("""
        INSERT INTO customers (id, nome, whatsapp_id, telefone, email, cpf_cnpj, endereco, cep,
            cidade, estado, data_ultima_compra, status_fidelidade, responsavel_id, origem, observacoes,
            recompra_dias, proxima_recompra, equipamentos, rolos_mes_media)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (cid, body["nome"], normalizar_telefone(body.get("whatsapp_id")), normalizar_telefone(body.get("telefone")), body.get("email"),
          cpf_cnpj_norm, body.get("endereco"), body.get("cep"), body.get("cidade"), body.get("estado"),
          body.get("data_ultima_compra"), body.get("status_fidelidade", "novo"),
          responsavel_id, body.get("origem"), body.get("observacoes"),
          recompra_dias, _ancora_recompra(body.get("data_ultima_compra"), recompra_dias),
          normalizar_equipamentos(body.get("equipamentos")),
          _int_or_none(body.get("rolos_mes_media"))))
    audit("create", "customers", cid, body)
    db.commit()
    return jsonify({"id": cid}), 201


@app.put("/api/customers/<customer_id>")
@login_required
def update_customer(customer_id):
    db = get_db()
    existing = db.execute("SELECT * FROM customers WHERE id = ?", (customer_id,)).fetchone()
    if not existing:
        return jsonify({"error": "Cliente não encontrado."}), 404
    if g.current_user["role"] != "admin" and existing["responsavel_id"] != g.current_user["id"]:
        return jsonify({"error": "Sem permissão para editar este cliente."}), 403
    body = request.get_json(force=True, silent=True) or {}
    if "cpf_cnpj" in body:
        doc_norm, erro_doc = validar_cpf_cnpj_ou_erro(body.get("cpf_cnpj"))
        if erro_doc:
            return jsonify({"error": erro_doc}), 400
        body["cpf_cnpj"] = doc_norm
    if "equipamentos" in body:
        body["equipamentos"] = normalizar_equipamentos(body.get("equipamentos"))
    for _tel in ("telefone", "whatsapp_id"):
        if _tel in body:
            body[_tel] = normalizar_telefone(body.get(_tel))
    if body.get("cpf_cnpj") or body.get("whatsapp_id"):
        erro_dup = _erro_cliente_duplicado(get_db(), body.get("cpf_cnpj"), body.get("whatsapp_id"),
                                           ignorar_id=customer_id)
        if erro_dup:
            return erro_dup
    if "rolos_mes_media" in body:
        body["rolos_mes_media"] = _int_or_none(body.get("rolos_mes_media"))
    campos = ["nome", "whatsapp_id", "telefone", "email", "cpf_cnpj", "endereco", "cep",
              "cidade", "estado", "data_ultima_compra", "status_fidelidade", "observacoes", "ativo",
              "equipamentos", "rolos_mes_media"]
    valores = {c: body.get(c, existing[c]) for c in campos}
    if "ativo" in body:
        valores["ativo"] = 1 if body.get("ativo") in (1, "1", True, "true") else 0
    dias_final = _int_or_none(body.get("recompra_dias")) if "recompra_dias" in body else existing["recompra_dias"]
    if "recompra_dias" in body or "data_ultima_compra" in body:
        # SEMPRE recalcula ao salvar a edição: próximo contato = última compra
        # + ciclo (ou hoje + ciclo sem última compra). Como a fórmula é
        # determinística, recalcular é idempotente — e conserta agendas
        # gravadas erradas por versões antigas com um simples "Salvar".
        # Desligar o ciclo (vazio/zero) limpa a agenda.
        db.execute("UPDATE customers SET recompra_dias = ?, proxima_recompra = ? WHERE id = ?",
                   (dias_final, _ancora_recompra(valores.get("data_ultima_compra"), dias_final), customer_id))
    db.execute(f"""
        UPDATE customers SET {", ".join(f"{c} = ?" for c in campos)}, updated_at = ?
        WHERE id = ?
    """, (*valores.values(), now_iso(), customer_id))
    audit("update", "customers", customer_id, body)
    db.commit()
    return jsonify({"ok": True})


@app.delete("/api/customers/<customer_id>")
@login_required
def delete_customer(customer_id):
    db = get_db()
    existing = db.execute("SELECT * FROM customers WHERE id = ?", (customer_id,)).fetchone()
    if not existing:
        return jsonify({"error": "Cliente não encontrado."}), 404
    if g.current_user["role"] != "admin" and existing["responsavel_id"] != g.current_user["id"]:
        return jsonify({"error": "Sem permissão para excluir este cliente."}), 403
    # Exclusão em cascata: negócios, compromissos e alertas ligados a este cliente somem
    # junto (definido via ON DELETE CASCADE no schema). Tarefas delegadas que citavam este
    # cliente são preservadas, só perdem o vínculo (ON DELETE SET NULL).
    db.execute("DELETE FROM customers WHERE id = ?", (customer_id,))
    audit("delete", "customers", customer_id, {"nome": existing["nome"]})
    db.commit()
    return jsonify({"ok": True})


# ------------------------------------------------------------------
# Rotas — Consulta externa de CNPJ e CEP (preenchimento automático)
# Usa BrasilAPI (dados da Receita Federal) e ViaCEP — ambas gratuitas,
# públicas e sem necessidade de chave de API. Nenhum custo envolvido.
# ------------------------------------------------------------------
def _http_get_json(url, timeout=6):
    """GET simples usando só a biblioteca padrão do Python (sem dependência
    nova). Levanta urllib.error.HTTPError em respostas 4xx/5xx."""
    req = urllib.request.Request(url, headers={
        "User-Agent": "CRM-Digimagem/1.0 (contato@lojadigimagem.com.br)",
        "Accept": "application/json",
    })
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


@app.get("/api/lookup/cnpj/<cnpj>")
@login_required
def lookup_cnpj(cnpj):
    # CNPJ alfanumérico: manter as letras! Normaliza (maiúsculas, sem máscara)
    # e valida o dígito verificador antes de gastar uma consulta externa.
    doc = normalizar_cpf_cnpj(cnpj)
    if len(doc) != 14 or not CNPJ_RE.match(doc):
        return jsonify({"error": "CNPJ precisa ter 14 caracteres (12 alfanuméricos + 2 dígitos verificadores)."}), 400
    if not cnpj_valido(doc):
        return jsonify({"error": "CNPJ inválido — o dígito verificador não confere."}), 400
    try:
        dados = _http_get_json(f"https://brasilapi.com.br/api/cnpj/v1/{doc}")
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return jsonify({"error": "CNPJ não encontrado na Receita Federal."}), 404
        return jsonify({"error": "Não foi possível consultar o CNPJ agora. Tente novamente ou preencha manualmente."}), 502
    except Exception:
        return jsonify({"error": "Não foi possível consultar o CNPJ agora. Tente novamente ou preencha manualmente."}), 502

    endereco_partes = [dados.get("logradouro"), dados.get("numero")]
    if dados.get("complemento"):
        endereco_partes.append(dados["complemento"])
    endereco = ", ".join(p for p in endereco_partes if p) or None

    telefone = dados.get("ddd_telefone_1") or ""

    return jsonify({
        "nome": dados.get("nome_fantasia") or dados.get("razao_social"),
        "razao_social": dados.get("razao_social"),
        "endereco": endereco,
        "bairro": dados.get("bairro"),
        "cep": dados.get("cep"),
        "cidade": dados.get("municipio"),
        "estado": dados.get("uf"),
        "telefone": telefone,
        "situacao_cadastral": dados.get("descricao_situacao_cadastral"),
    })


@app.get("/api/lookup/cep/<cep>")
@login_required
def lookup_cep(cep):
    digitos = "".join(c for c in cep if c.isdigit())
    if len(digitos) != 8:
        return jsonify({"error": "CEP precisa ter 8 dígitos."}), 400
    try:
        dados = _http_get_json(f"https://viacep.com.br/ws/{digitos}/json/")
    except Exception:
        return jsonify({"error": "Não foi possível consultar o CEP agora. Tente novamente ou preencha manualmente."}), 502

    if dados.get("erro"):
        return jsonify({"error": "CEP não encontrado."}), 404

    endereco_partes = [dados.get("logradouro"), dados.get("bairro")]
    endereco = ", ".join(p for p in endereco_partes if p) or None

    return jsonify({
        "endereco": endereco,
        "cidade": dados.get("localidade"),
        "estado": dados.get("uf"),
    })


# ------------------------------------------------------------------
# Rotas — Deals / Funil
# ------------------------------------------------------------------
ETAPAS = ["novo_lead", "qualificacao", "proposta_enviada", "negociacao", "fechado_ganho", "fechado_perdido"]
ETAPAS_ABERTAS = ["novo_lead", "qualificacao", "proposta_enviada", "negociacao"]
ETAPA_LABEL_BACKEND = {
    "novo_lead": "Novo Lead", "qualificacao": "Qualificação",
    "proposta_enviada": "Proposta Enviada", "negociacao": "Negociação",
}


@app.get("/api/deals")
@login_required
def list_deals():
    db = get_db()
    clause, params = scope_filter_clause("user_id")
    rows = db.execute(f"""
        SELECT d.*, c.nome as cliente_nome, u.nome as vendedor_nome,
               (SELECT COUNT(*) FROM deal_notes n WHERE n.deal_id = d.id) as notas_count
        FROM deals d
        JOIN customers c ON c.id = d.customer_id
        LEFT JOIN users u ON u.id = d.user_id
        WHERE 1=1 {clause}
        ORDER BY d.created_at DESC
    """, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/deals")
@login_required
def create_deal():
    body = request.get_json(force=True, silent=True) or {}
    for f in ("customer_id", "titulo"):
        if not body.get(f):
            return jsonify({"error": f"Campo '{f}' é obrigatório."}), 400
    db = get_db()

    cliente = db.execute("SELECT * FROM customers WHERE id = ?", (body["customer_id"],)).fetchone()
    if not cliente:
        return jsonify({"error": "Cliente não encontrado."}), 404
    if g.current_user["role"] != "admin" and cliente["responsavel_id"] != g.current_user["id"]:
        return jsonify({"error": "Você só pode criar negócios para os seus próprios clientes."}), 403

    # Vendedor só pode criar negócio para si mesmo; só admin pode atribuir a outro usuário.
    user_id = body.get("user_id", g.current_user["id"])
    if g.current_user["role"] != "admin":
        user_id = g.current_user["id"]

    categoria = body.get("categoria") or "padrao"
    if categoria not in ("padrao", "software"):
        return jsonify({"error": "Categoria de negócio inválida."}), 400
    produto_software = None
    if categoria == "software":
        produto_software = body.get("produto_software")
        if produto_software not in PRODUTOS_SOFTWARE:
            return jsonify({"error": "Escolha o produto: Revele Momentos ou Revele Momentos Frontier."}), 400
    # 📦 produto do catálogo (opcional; só para vendas padrão): cria o elo
    # estruturado com as metas por produto, sem tirar a venda dos relatórios.
    produto_id = None
    if categoria == "padrao" and body.get("produto_id"):
        pr = db.execute("SELECT id FROM produtos WHERE id = ? AND ativo = 1",
                        (body.get("produto_id"),)).fetchone()
        if not pr:
            return jsonify({"error": "Produto do catálogo inválido ou inativo."}), 400
        produto_id = pr["id"]
    produto_qtd = (_int_or_none(body.get("produto_qtd")) or 1) if produto_id else None
    # 🧾 itens do orçamento na criação — validação COMPLETA antes de gravar
    itens_norm = []
    if body.get("itens"):
        if categoria == "software":
            return jsonify({"error": "Itens de orçamento são para venda padrão — o software é uma assinatura, sem itens."}), 400
        if not isinstance(body["itens"], list):
            return jsonify({"error": "Formato inválido dos itens do orçamento."}), 400
        eh_admin = g.current_user["role"] == "admin"
        for idx, it in enumerate(body["itens"], 1):
            p = db.execute("SELECT * FROM produtos WHERE id = ? AND ativo = 1 AND ofertavel != 0",
                           ((it or {}).get("produto_id"),)).fetchone()
            if not p:
                return jsonify({"error": f"Item {idx}: produto inválido ou fora de oferta."}), 400
            qtd = _int_or_none(it.get("qtd"))
            if not qtd or qtd < 1:
                return jsonify({"error": f"Item {idx} ({p['nome']}): informe a quantidade."}), 400
            try:
                preco = round(float(it.get("preco_unit")), 2) if it.get("preco_unit") not in (None, "") else None
            except (TypeError, ValueError):
                preco = None
            if preco is None:
                preco = p["preco_tabela"]
            if preco is None or preco <= 0:
                return jsonify({"error": f"Item {idx} ({p['nome']}): informe o preço unitário."}), 400
            limite = p["preco_limite"]
            abaixo_limite = bool(limite and preco < limite and not eh_admin)
            usou_limite = 1 if (p["preco_tabela"] and preco < p["preco_tabela"]) else 0
            itens_norm.append({"produto_id": p["id"], "nome": p["nome"], "qtd": qtd,
                               "preco": preco, "usou_limite": usou_limite,
                               "abaixo_limite": abaixo_limite})

    did = new_id()
    etapa_inicial = body.get("etapa_funil", "novo_lead")
    db.execute("""
        INSERT INTO deals (id, customer_id, user_id, titulo, etapa_funil, valor_estimado, status,
                           data_prevista_fechamento, categoria, produto_software, produto_id, produto_qtd)
        VALUES (?,?,?,?,?,?,'aberto',?,?,?,?,?)
    """, (did, body["customer_id"], user_id, body["titulo"],
          etapa_inicial, body.get("valor_estimado", 0), body.get("data_prevista_fechamento"),
          categoria, produto_software, produto_id, produto_qtd))
    if itens_norm:
        for it in itens_norm:
            iid = new_id()
            lib_id = None
            if it["abaixo_limite"]:
                lib_id = new_id()
                db.execute("""
                    INSERT INTO liberacoes_preco (id, deal_id, produto_id, user_id, preco_pedido, motivo, created_at)
                    VALUES (?,?,?,?,?,?,?)
                """, (lib_id, did, it["produto_id"], g.current_user["id"], it["preco"],
                      "Pedido gerado na criação do negócio", now_iso()))
            db.execute("""
                INSERT INTO deal_itens (id, deal_id, produto_id, qtd, preco_unit, usou_limite,
                                        aprovado, liberacao_id, user_id, created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (iid, did, it["produto_id"], it["qtd"], it["preco"], it["usou_limite"],
                  0 if it["abaixo_limite"] else 1, lib_id, g.current_user["id"], now_iso()))
            audit("create", "deal_itens", iid, {"deal": did, "produto": it["nome"], "qtd": it["qtd"],
                                                "preco": it["preco"], "abaixo_tabela": bool(it["usou_limite"]),
                                                "abaixo_limite": it["abaixo_limite"]})
        subtotal = round(sum(it["preco"] * it["qtd"] for it in itens_norm), 2)
        db.execute("UPDATE deals SET valor_estimado = ? WHERE id = ?", (subtotal, did))
    # 💰 valor inicial entra no histórico (valor_anterior NULL = "inicial")
    db.execute("""
        INSERT INTO deal_value_history (id, deal_id, valor_anterior, valor_novo, user_id, created_at)
        VALUES (?,?,NULL,?,?,?)
    """, (new_id(), did, body.get("valor_estimado", 0) or 0, user_id, now_iso()))
    db.execute("""
        INSERT INTO deal_stage_history (id, deal_id, etapa_anterior, etapa_nova, user_id, data_transicao)
        VALUES (?,?,NULL,?,?,?)
    """, (new_id(), did, etapa_inicial, user_id, now_iso()))
    audit("create", "deals", did, body)
    db.commit()
    return jsonify({"id": did}), 201


MOTIVOS_PERDA_VALIDOS = (
    "preco", "concorrente", "falta_funcionalidade", "sumiu_no_show",
    "sem_orcamento", "timing_errado", "outro",
)


@app.post("/api/deals/<deal_id>/stage")
@login_required
def move_deal_stage(deal_id):
    db = get_db()
    existing = db.execute("SELECT * FROM deals WHERE id = ?", (deal_id,)).fetchone()
    if not existing:
        return jsonify({"error": "Negócio não encontrado."}), 404
    if g.current_user["role"] != "admin" and existing["user_id"] != g.current_user["id"]:
        return jsonify({"error": "Sem permissão para alterar este negócio."}), 403
    body = request.get_json(force=True, silent=True) or {}
    nova_etapa = body.get("etapa_funil")
    if nova_etapa not in ETAPAS:
        return jsonify({"error": "Etapa inválida."}), 400

    status = "aberto"
    motivo_perda = existing["motivo_perda"]
    motivo_perda_detalhe = existing["motivo_perda_detalhe"]
    if nova_etapa not in ("fechado_ganho", "fechado_perdido"):
        # Movimentação para etapa aberta (inclui reabertura): o motivo da
        # perda sai do negócio — ele permanece registrado no marco "Perdido"
        # da linha do tempo (deal_stage_history.motivo_perda).
        motivo_perda, motivo_perda_detalhe = None, None
    if nova_etapa == "fechado_ganho":
        # 🛡 trava de faturamento: itens abaixo do limite precisam de liberação
        pendentes = db.execute("""
            SELECT p.nome, i.preco_unit, p.preco_limite FROM deal_itens i
            JOIN produtos p ON p.id = i.produto_id
            WHERE i.deal_id = ? AND i.aprovado = 0
        """, (deal_id,)).fetchall()
        if pendentes:
            nomes = "; ".join(f"{r['nome']} a R$ {r['preco_unit']:.2f} (limite R$ {r['preco_limite']:.2f})"
                              for r in pendentes[:3])
            return jsonify({"error": f"Não é possível faturar: {len(pendentes)} item(ns) do orçamento aguardam liberação do administrador — {nomes}."}), 400
        status = "ganho"
        motivo_perda, motivo_perda_detalhe = None, None
    elif nova_etapa == "fechado_perdido":
        status = "perdido"
        motivo_perda = body.get("motivo_perda")
        if motivo_perda not in MOTIVOS_PERDA_VALIDOS:
            return jsonify({"error": "Informe o motivo da perda para registrar este negócio como perdido."}), 400
        motivo_perda_detalhe = body.get("motivo_perda_detalhe")

    db.execute("""
        UPDATE deals SET etapa_funil = ?, status = ?, etapa_atualizada_em = ?, updated_at = ?,
            motivo_perda = ?, motivo_perda_detalhe = ?
        WHERE id = ?
    """, (nova_etapa, status, now_iso(), now_iso(), motivo_perda, motivo_perda_detalhe, deal_id))

    db.execute("""
        INSERT INTO deal_stage_history (id, deal_id, etapa_anterior, etapa_nova, user_id,
                                        motivo_perda, motivo_perda_detalhe, data_transicao)
        VALUES (?,?,?,?,?,?,?,?)
    """, (new_id(), deal_id, existing["etapa_funil"], nova_etapa, g.current_user["id"],
          motivo_perda if nova_etapa == "fechado_perdido" else None,
          motivo_perda_detalhe if nova_etapa == "fechado_perdido" else None,
          now_iso()))

    if nova_etapa == "fechado_ganho":
        # Venda realizada: atualiza a última compra do cliente (a régua de
        # inatividade zera na raiz) e resolve automaticamente os alertas
        # ativos dele no feed — o aviso some sozinho, como deve ser.
        db.execute("UPDATE customers SET data_ultima_compra = ? WHERE id = ?",
                   (datetime.now(timezone.utc).strftime("%Y-%m-%d"), existing["customer_id"]))
        db.execute("""
            UPDATE ai_insights SET lido = 1
            WHERE customer_id = ? AND lido = 0
              AND tipo_alerta IN ('inatividade', 'abandono_funil', 'sugestao_conteudo')
        """, (existing["customer_id"],))

    # 🔁 Recompra programada: ao FECHAR o negócio (ganho ou perdido), agenda o
    # próximo contato do cliente para daqui a N dias. O motor de alertas criará
    # o novo negócio e o compromisso quando a data chegar.
    if nova_etapa in ("fechado_ganho", "fechado_perdido"):
        cli = db.execute("SELECT recompra_dias FROM customers WHERE id = ?",
                         (existing["customer_id"],)).fetchone()
        if cli and cli["recompra_dias"]:
            db.execute("UPDATE customers SET proxima_recompra = ? WHERE id = ?",
                       (_hoje_mais_dias(cli["recompra_dias"]), existing["customer_id"]))

    audit("update", "deals", deal_id, {"nova_etapa": nova_etapa, "motivo_perda": motivo_perda})
    db.commit()
    return jsonify({"ok": True})


@app.post("/api/deals/<deal_id>/reabrir")
@login_required
def reopen_deal(deal_id):
    """Reabre um negócio marcado como perdido: ele volta ao funil na etapa
    em que estava antes de ser perdido (fallback: negociação). O motivo da
    perda continua registrado no marco 'Perdido' da linha do tempo — nada
    da história se perde. Permissão: dono do negócio ou admin."""
    db = get_db()
    existing = db.execute("SELECT * FROM deals WHERE id = ?", (deal_id,)).fetchone()
    if not existing:
        return jsonify({"error": "Negócio não encontrado."}), 404
    if g.current_user["role"] != "admin" and existing["user_id"] != g.current_user["id"]:
        return jsonify({"error": "Sem permissão para alterar este negócio."}), 403
    if existing["status"] != "perdido":
        return jsonify({"error": "Só é possível reabrir negócios marcados como perdidos."}), 400

    ultima_perda = db.execute("""
        SELECT * FROM deal_stage_history
        WHERE deal_id = ? AND etapa_nova = 'fechado_perdido'
        ORDER BY data_transicao DESC, rowid DESC LIMIT 1
    """, (deal_id,)).fetchone()

    # Volta para a etapa em que o negócio estava quando foi perdido
    etapa_retorno = "negociacao"
    if ultima_perda and ultima_perda["etapa_anterior"] in ETAPAS             and ultima_perda["etapa_anterior"] not in ("fechado_ganho", "fechado_perdido"):
        etapa_retorno = ultima_perda["etapa_anterior"]

    # Perdas registradas antes do motivo ser gravado no marco: preserva o
    # motivo no próprio marco antes de limpá-lo do negócio.
    if ultima_perda and not ultima_perda["motivo_perda"] and existing["motivo_perda"]:
        db.execute("""UPDATE deal_stage_history SET motivo_perda = ?, motivo_perda_detalhe = ?
                      WHERE id = ?""",
                   (existing["motivo_perda"], existing["motivo_perda_detalhe"], ultima_perda["id"]))

    db.execute("""
        UPDATE deals SET etapa_funil = ?, status = 'aberto', etapa_atualizada_em = ?, updated_at = ?,
            motivo_perda = NULL, motivo_perda_detalhe = NULL
        WHERE id = ?
    """, (etapa_retorno, now_iso(), now_iso(), deal_id))
    db.execute("""
        INSERT INTO deal_stage_history (id, deal_id, etapa_anterior, etapa_nova, user_id, data_transicao)
        VALUES (?,?,?,?,?,?)
    """, (new_id(), deal_id, "fechado_perdido", etapa_retorno, g.current_user["id"], now_iso()))
    audit("update", "deals", deal_id, {"acao": "reabrir", "etapa_retorno": etapa_retorno})
    db.commit()
    return jsonify({"ok": True, "etapa_funil": etapa_retorno})


@app.get("/api/deals/<deal_id>")
@login_required
def get_deal(deal_id):
    db = get_db()
    d = db.execute("""
        SELECT deals.*, customers.nome as cliente_nome FROM deals
        JOIN customers ON customers.id = deals.customer_id
        WHERE deals.id = ?
    """, (deal_id,)).fetchone()
    if not d:
        return jsonify({"error": "Negócio não encontrado."}), 404
    if g.current_user["role"] != "admin" and d["user_id"] != g.current_user["id"]:
        return jsonify({"error": "Sem permissão para ver este negócio."}), 403
    return jsonify(dict(d))


# ------------------------------------------------------------------
# Rotas — Histórico da Negociação (conversas + mudanças de etapa)
# Regra de visibilidade: vendedor só acessa o histórico dos PRÓPRIOS
# negócios; admin acessa o histórico de qualquer negócio, inclusive
# os de outros administradores. As notas são imutáveis (sem edição ou
# exclusão) para o histórico ser um registro confiável do que ocorreu.
# ------------------------------------------------------------------
# ------------------------------------------------------------------
# 🧾 ORÇAMENTO DO NEGÓCIO (Fase 2) — o simulador nativo
# ------------------------------------------------------------------
def _calcular_orcamento(db, d):
    """Monta o orçamento completo do negócio: itens, subtotal, acréscimo da
    condição de pagamento, total e situação do frete pela UF do cliente."""
    itens = [dict(r) for r in db.execute("""
        SELECT i.*, p.nome as produto_nome, p.embalagem, p.preco_tabela, p.preco_limite,
               l.status as liberacao_status, l.observacao as liberacao_obs
        FROM deal_itens i JOIN produtos p ON p.id = i.produto_id
        LEFT JOIN liberacoes_preco l ON l.id = i.liberacao_id
        WHERE i.deal_id = ? ORDER BY i.created_at
    """, (d["id"],)).fetchall()]
    subtotal = sum(i["preco_unit"] * i["qtd"] for i in itens)
    condicao = None
    acrescimo = 0.0
    if d["condicao_pagamento_id"]:
        c = db.execute("SELECT * FROM condicoes_pagamento WHERE id = ?",
                       (d["condicao_pagamento_id"],)).fetchone()
        if c:
            condicao = dict(c)
            acrescimo = round(subtotal * (c["acrescimo_pct"] or 0) / 100, 2)
    total = round(subtotal + acrescimo, 2)
    cliente = db.execute("SELECT estado FROM customers WHERE id = ?", (d["customer_id"],)).fetchone()
    uf = (cliente["estado"] or "").strip().upper() if cliente else ""
    frete = None
    if uf:
        f = db.execute("SELECT * FROM frete_uf WHERE uf = ?", (uf,)).fetchone()
        if f:
            falta = max(0.0, round(f["minimo"] - total, 2))
            frete = {"uf": uf, "minimo": f["minimo"], "falta": falta, "gratis": falta <= 0}
    return {"itens": itens, "subtotal": round(subtotal, 2), "condicao": condicao,
            "acrescimo": acrescimo, "total": total, "frete": frete,
            "pendentes_liberacao": sum(1 for i in itens if not i["aprovado"])}


def _sincronizar_valor_negocio(db, deal_id):
    """O valor do negócio acompanha o total do orçamento (sem marcos no
    histórico de valores — o orçamento em si já é o detalhamento)."""
    d = db.execute("SELECT * FROM deals WHERE id = ?", (deal_id,)).fetchone()
    calc = _calcular_orcamento(db, d)
    if calc["itens"]:
        db.execute("UPDATE deals SET valor_estimado = ?, updated_at = ? WHERE id = ?",
                   (calc["total"], now_iso(), deal_id))
    return calc


def _liberacao_vigente(db, deal_id, produto_id, user_id):
    """Liberação aprovada, não usada e dentro da validade (7 dias)."""
    limite_data = (datetime.now(timezone.utc) - timedelta(days=7)).strftime("%Y-%m-%d %H:%M:%S")
    return db.execute("""
        SELECT * FROM liberacoes_preco
        WHERE deal_id = ? AND produto_id = ? AND user_id = ?
          AND status = 'aprovada' AND created_at >= ?
        ORDER BY created_at DESC LIMIT 1
    """, (deal_id, produto_id, user_id, limite_data)).fetchone()


@app.post("/api/deals/<deal_id>/liberacoes")
@login_required
def pedir_liberacao(deal_id):
    """🙋 Vendedor pede liberação para um preço abaixo do limite."""
    d, erro = _deal_permitido_ou_erro(deal_id, escrita=True)
    if erro:
        return erro
    body = request.get_json(force=True, silent=True) or {}
    db = get_db()
    p = db.execute("SELECT * FROM produtos WHERE id = ? AND ativo = 1 AND ofertavel != 0",
                   (body.get("produto_id"),)).fetchone()
    if not p:
        return jsonify({"error": "Produto inválido."}), 400
    try:
        preco = round(float(body.get("preco_pedido")), 2)
    except (TypeError, ValueError):
        preco = None
    if not preco or preco <= 0:
        return jsonify({"error": "Informe o preço pretendido."}), 400
    if p["preco_limite"] and preco >= p["preco_limite"]:
        return jsonify({"error": "Este preço já está dentro da sua autonomia — pode lançar direto no orçamento."}), 400
    ja = db.execute("""
        SELECT id FROM liberacoes_preco WHERE deal_id = ? AND produto_id = ? AND user_id = ? AND status = 'pendente'
    """, (deal_id, p["id"], g.current_user["id"])).fetchone()
    if ja:
        return jsonify({"error": "Já existe um pedido pendente para este produto neste negócio."}), 400
    lid = new_id()
    db.execute("""
        INSERT INTO liberacoes_preco (id, deal_id, produto_id, user_id, preco_pedido, motivo, created_at)
        VALUES (?,?,?,?,?,?,?)
    """, (lid, deal_id, p["id"], g.current_user["id"], preco,
          (body.get("motivo") or "").strip()[:400] or None, now_iso()))
    audit("create", "liberacoes_preco", lid, {"deal": deal_id, "produto": p["nome"], "preco": preco})
    db.commit()
    return jsonify({"ok": True, "id": lid}), 201


@app.get("/api/liberacoes")
@login_required
def listar_liberacoes():
    """Admin: todos os pedidos (pendentes primeiro). Vendedor: só os seus."""
    db = get_db()
    clause, params = ("", [])
    if g.current_user["role"] != "admin":
        clause, params = (" WHERE l.user_id = ?", [g.current_user["id"]])
    rows = db.execute(f"""
        SELECT l.*, p.nome as produto_nome, p.preco_tabela, p.preco_limite,
               u.nome as vendedor_nome, a.nome as admin_nome,
               d.titulo as deal_titulo, c.nome as cliente_nome
        FROM liberacoes_preco l
        JOIN produtos p ON p.id = l.produto_id
        JOIN users u ON u.id = l.user_id
        LEFT JOIN users a ON a.id = l.admin_id
        JOIN deals d ON d.id = l.deal_id
        JOIN customers c ON c.id = d.customer_id
        {clause}
        ORDER BY CASE l.status WHEN 'pendente' THEN 0 ELSE 1 END, l.created_at DESC
        LIMIT 100
    """, params).fetchall()
    return jsonify({"liberacoes": [dict(r) for r in rows], "admin": g.current_user["role"] == "admin"})


@app.post("/api/liberacoes/<liberacao_id>/decidir")
@login_required
def decidir_liberacao(liberacao_id):
    """Admin aprova (podendo ajustar o preço) ou nega o pedido."""
    if g.current_user["role"] != "admin":
        return jsonify({"error": "Apenas administradores decidem liberações de preço."}), 403
    db = get_db()
    l = db.execute("SELECT * FROM liberacoes_preco WHERE id = ?", (liberacao_id,)).fetchone()
    if not l:
        return jsonify({"error": "Pedido não encontrado."}), 404
    if l["status"] != "pendente":
        return jsonify({"error": f"Este pedido já foi {l['status']}."}), 400
    body = request.get_json(force=True, silent=True) or {}
    decisao = body.get("decisao")
    if decisao not in ("aprovar", "negar"):
        return jsonify({"error": "Decisão inválida."}), 400
    obs = (body.get("observacao") or "").strip()[:400] or None
    if decisao == "negar":
        db.execute("""UPDATE liberacoes_preco SET status = 'negada', admin_id = ?, observacao = ?,
                      decidido_em = ? WHERE id = ?""",
                   (g.current_user["id"], obs, now_iso(), liberacao_id))
    else:
        try:
            autorizado = round(float(body.get("preco_autorizado", l["preco_pedido"])), 2)
        except (TypeError, ValueError):
            autorizado = l["preco_pedido"]
        if autorizado <= 0:
            return jsonify({"error": "Preço autorizado inválido."}), 400
        db.execute("""UPDATE liberacoes_preco SET status = 'aprovada', preco_autorizado = ?,
                      admin_id = ?, observacao = ?, decidido_em = ? WHERE id = ?""",
                   (autorizado, g.current_user["id"], obs, now_iso(), liberacao_id))
        # libera o item do orçamento (e ajusta o preço, se o admin mudou)
        db.execute("""UPDATE deal_itens SET aprovado = 1, preco_unit = ?
                      WHERE liberacao_id = ?""", (autorizado, liberacao_id))
        _sincronizar_valor_negocio(db, l["deal_id"])
    audit("update", "liberacoes_preco", liberacao_id, {"decisao": decisao})
    db.commit()
    return jsonify({"ok": True})


@app.get("/api/deals/<deal_id>/orcamento")
@login_required
def ver_orcamento(deal_id):
    d, erro = _deal_permitido_ou_erro(deal_id)
    if erro:
        return erro
    db = get_db()
    calc = _calcular_orcamento(db, d)
    condicoes = [dict(r) for r in db.execute(
        "SELECT * FROM condicoes_pagamento WHERE simples = 1 ORDER BY ordem").fetchall()]
    produtos = [dict(r) for r in db.execute("""
        SELECT id, nome, embalagem, preco_tabela, preco_limite FROM produtos
        WHERE ativo = 1 AND ofertavel != 0 ORDER BY nome
    """).fetchall()]
    if g.current_user["role"] != "admin":
        for p in produtos:
            pass  # vendedor precisa do limite para a trava avisar ANTES? Não: o servidor valida; o campo não é enviado
        produtos = [{k: v for k, v in p.items() if k != "preco_limite"} for p in produtos]
    return jsonify({**calc, "condicoes_disponiveis": condicoes, "produtos_disponiveis": produtos,
                    "pode_editar": g.current_user["role"] == "admin" or d["user_id"] == g.current_user["id"],
                    "admin": g.current_user["role"] == "admin"})


@app.post("/api/deals/<deal_id>/orcamento/itens")
@login_required
def add_item_orcamento(deal_id):
    d, erro = _deal_permitido_ou_erro(deal_id, escrita=True)
    if erro:
        return erro
    if d["status"] != "aberto":
        return jsonify({"error": "Orçamento só pode ser alterado em negócio ABERTO — reabra o negócio se necessário."}), 400
    body = request.get_json(force=True, silent=True) or {}
    db = get_db()
    p = db.execute("SELECT * FROM produtos WHERE id = ? AND ativo = 1 AND ofertavel != 0",
                   (body.get("produto_id"),)).fetchone()
    if not p:
        return jsonify({"error": "Escolha um produto válido do catálogo (ativo e ofertável)."}), 400
    qtd = _int_or_none(body.get("qtd"))
    if not qtd or qtd < 1:
        return jsonify({"error": "Informe a quantidade (número maior que zero)."}), 400
    preco_unit = body.get("preco_unit")
    try:
        preco_unit = round(float(preco_unit), 2)
    except (TypeError, ValueError):
        preco_unit = None
    if preco_unit is None:
        preco_unit = p["preco_tabela"]
    if preco_unit is None or preco_unit <= 0:
        return jsonify({"error": "Informe o preço unitário (o produto não tem preço de tabela cadastrado)."}), 400
    # 🛡 TRAVA DO PREÇO-LIMITE
    limite = p["preco_limite"]
    liberacao = None
    aprovado = 1
    lib_id = None
    if limite and preco_unit < limite and g.current_user["role"] != "admin":
        # 🙋 abaixo do limite: SALVA assim mesmo (a proposta é negociação),
        # porém pendente de liberação — e o pedido vai sozinho ao admin.
        liberacao = _liberacao_vigente(db, deal_id, p["id"], g.current_user["id"])
        if liberacao and preco_unit >= liberacao["preco_autorizado"]:
            aprovado = 1  # já havia liberação vigente que cobre este preço
        else:
            aprovado = 0
            lib_id = new_id()
            db.execute("""
                INSERT INTO liberacoes_preco (id, deal_id, produto_id, user_id, preco_pedido, motivo, created_at)
                VALUES (?,?,?,?,?,?,?)
            """, (lib_id, deal_id, p["id"], g.current_user["id"], preco_unit,
                  (body.get("motivo") or "").strip()[:400] or None, now_iso()))
            liberacao = None
    usou_limite = 1 if (p["preco_tabela"] and preco_unit < p["preco_tabela"]) else 0
    iid = new_id()
    db.execute("""
        INSERT INTO deal_itens (id, deal_id, produto_id, qtd, preco_unit, usou_limite, aprovado,
                                liberacao_id, user_id, created_at)
        VALUES (?,?,?,?,?,?,?,?,?,?)
    """, (iid, deal_id, p["id"], qtd, preco_unit, usou_limite, aprovado, lib_id,
          g.current_user["id"], now_iso()))
    audit("create", "deal_itens", iid, {"deal": deal_id, "produto": p["nome"], "qtd": qtd,
                                        "preco": preco_unit, "abaixo_tabela": bool(usou_limite),
                                        "abaixo_limite": bool(limite and preco_unit < limite),
                                        "liberacao": liberacao["id"] if liberacao else None})
    if liberacao:
        db.execute("UPDATE liberacoes_preco SET status = 'usada', usado_em = ? WHERE id = ?",
                   (now_iso(), liberacao["id"]))
    calc = _sincronizar_valor_negocio(db, deal_id)
    db.commit()
    return jsonify({"ok": True, "id": iid, "total": calc["total"]}), 201


@app.delete("/api/deals/<deal_id>/orcamento/itens/<item_id>")
@login_required
def remover_item_orcamento(deal_id, item_id):
    d, erro = _deal_permitido_ou_erro(deal_id, escrita=True)
    if erro:
        return erro
    if d["status"] != "aberto":
        return jsonify({"error": "Orçamento só pode ser alterado em negócio ABERTO."}), 400
    db = get_db()
    i = db.execute("SELECT * FROM deal_itens WHERE id = ? AND deal_id = ?", (item_id, deal_id)).fetchone()
    if not i:
        return jsonify({"error": "Item não encontrado neste orçamento."}), 404
    db.execute("DELETE FROM deal_itens WHERE id = ?", (item_id,))
    audit("delete", "deal_itens", item_id, {"deal": deal_id})
    _sincronizar_valor_negocio(db, deal_id)
    db.commit()
    return jsonify({"ok": True})


@app.put("/api/deals/<deal_id>/orcamento/condicao")
@login_required
def definir_condicao_orcamento(deal_id):
    d, erro = _deal_permitido_ou_erro(deal_id, escrita=True)
    if erro:
        return erro
    body = request.get_json(force=True, silent=True) or {}
    db = get_db()
    cid = body.get("condicao_pagamento_id") or None
    if cid:
        c = db.execute("SELECT id FROM condicoes_pagamento WHERE id = ? AND eh_nota = 0", (cid,)).fetchone()
        if not c:
            return jsonify({"error": "Condição de pagamento inválida."}), 400
    db.execute("UPDATE deals SET condicao_pagamento_id = ?, updated_at = ? WHERE id = ?",
               (cid, now_iso(), deal_id))
    _sincronizar_valor_negocio(db, deal_id)
    db.commit()
    return jsonify({"ok": True})


@app.get("/api/deals/<deal_id>/proposta")
@login_required
def gerar_proposta(deal_id):
    """📄 Texto da proposta para o CLIENTE: itens, totais, pagamento e frete.
    NUNCA inclui preço-limite ou dados internos."""
    d, erro = _deal_permitido_ou_erro(deal_id)
    if erro:
        return erro
    db = get_db()
    calc = _calcular_orcamento(db, d)
    if not calc["itens"]:
        return jsonify({"error": "Adicione itens ao orçamento antes de gerar a proposta."}), 400
    fmt = lambda v: f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    linhas = [f"*Proposta Digimagem — {d['cliente_nome']}*", ""]
    for i in calc["itens"]:
        linhas.append(f"• {i['produto_nome']} — {i['qtd']}x {fmt(i['preco_unit'])} = {fmt(i['preco_unit'] * i['qtd'])}")
        if i["embalagem"]:
            linhas.append(f"   ({i['embalagem']})")
    linhas.append("")
    linhas.append(f"Subtotal: {fmt(calc['subtotal'])}")
    if calc["acrescimo"]:
        linhas.append(f"Acréscimo do cartão: {fmt(calc['acrescimo'])}")
    linhas.append(f"*Total: {fmt(calc['total'])}*")
    if calc["condicao"]:
        linhas.append(f"Pagamento: {calc['condicao']['forma']} — {calc['condicao']['condicao']}")
    if calc["frete"]:
        if calc["frete"]["gratis"]:
            linhas.append(f"🚚 *Frete GRÁTIS* para {calc['frete']['uf']} (pedido acima de {fmt(calc['frete']['minimo'])})")
        else:
            linhas.append(f"🚚 Faltam apenas {fmt(calc['frete']['falta'])} para o seu pedido ter *frete grátis* ({calc['frete']['uf']}: mínimo {fmt(calc['frete']['minimo'])})")
    linhas.append("")
    linhas.append("Preços vigentes até comunicado da Fujifilm. Fico à disposição!")
    return jsonify({"texto": "\n".join(linhas)})


@app.post("/api/customers/<customer_id>/transferir")
@login_required
def transferir_cliente(customer_id):
    """⇄ Transferência de titularidade (somente admin). Negócios e valores
    do dono anterior permanecem com ele (relatórios intactos); o novo dono
    ganha LEITURA das conversas antigas do cliente e começa negociações do
    zero. O registro guarda quem transferiu, quando, de quem e para quem —
    origem e destino visíveis apenas ao administrador."""
    if g.current_user["role"] != "admin":
        return jsonify({"error": "Apenas administradores transferem clientes."}), 403
    db = get_db()
    c = db.execute("SELECT * FROM customers WHERE id = ?", (customer_id,)).fetchone()
    if not c:
        return jsonify({"error": "Cliente não encontrado."}), 404
    body = request.get_json(force=True, silent=True) or {}
    novo_id = body.get("novo_responsavel_id")
    novo = db.execute("SELECT * FROM users WHERE id = ? AND ativo = 1", (novo_id,)).fetchone()
    if not novo:
        return jsonify({"error": "Escolha um usuário ativo para receber o cliente."}), 400
    if novo_id == c["responsavel_id"]:
        return jsonify({"error": "Este usuário já é o responsável por este cliente."}), 400
    db.execute("UPDATE customers SET responsavel_id = ? WHERE id = ?", (novo_id, customer_id))
    db.execute("""
        INSERT INTO customer_transfers (id, customer_id, de_user_id, para_user_id, admin_id, created_at)
        VALUES (?,?,?,?,?,?)
    """, (new_id(), customer_id, c["responsavel_id"], novo_id, g.current_user["id"], now_iso()))
    abertos = db.execute("SELECT COUNT(*) n FROM deals WHERE customer_id = ? AND status = 'aberto'",
                         (customer_id,)).fetchone()["n"]
    audit("update", "customers", customer_id, {"acao": "transferir", "para": novo_id})
    db.commit()
    return jsonify({"ok": True, "negocios_abertos": abertos})


def _deal_permitido_ou_erro(deal_id, escrita=False):
    """Carrega o negócio e aplica a regra de visibilidade:
    - admin: tudo;
    - dono do NEGÓCIO: lê e escreve;
    - dono ATUAL do CLIENTE (ex.: recebeu por transferência ⇄): apenas LÊ o
      histórico das negociações antigas — a "cópia" do que já foi conversado."""
    db = get_db()
    d = db.execute("""
        SELECT d.*, c.nome as cliente_nome, c.rolos_mes_media as cliente_rolos_mes,
               c.responsavel_id as cliente_responsavel_id,
               u.nome as vendedor_nome, pr.nome as produto_nome
        FROM deals d
        JOIN customers c ON c.id = d.customer_id
        LEFT JOIN users u ON u.id = d.user_id
        LEFT JOIN produtos pr ON pr.id = d.produto_id
        WHERE d.id = ?
    """, (deal_id,)).fetchone()
    if not d:
        return None, (jsonify({"error": "Negócio não encontrado."}), 404)
    if g.current_user["role"] != "admin":
        dono_negocio = d["user_id"] == g.current_user["id"]
        dono_cliente = d["cliente_responsavel_id"] == g.current_user["id"]
        if escrita and not dono_negocio:
            return None, (jsonify({"error": "Somente o vendedor deste negócio (ou um administrador) pode escrever no histórico."}), 403)
        if not (dono_negocio or dono_cliente):
            return None, (jsonify({"error": "Sem permissão para acessar o histórico deste negócio."}), 403)
    return d, None


@app.get("/api/deals/<deal_id>/historico")
@login_required
def deal_historico(deal_id):
    d, erro = _deal_permitido_ou_erro(deal_id)
    if erro:
        return erro
    db = get_db()
    notas = db.execute("""
        SELECT n.*, u.nome as autor_nome, u.role as autor_role FROM deal_notes n
        LEFT JOIN users u ON u.id = n.user_id
        WHERE n.deal_id = ?
    """, (deal_id,)).fetchall()
    etapas = db.execute("""
        SELECT h.*, u.nome as autor_nome FROM deal_stage_history h
        LEFT JOIN users u ON u.id = h.user_id
        WHERE h.deal_id = ?
    """, (deal_id,)).fetchall()
    valores = db.execute("""
        SELECT v.*, u.nome as autor_nome FROM deal_value_history v
        LEFT JOIN users u ON u.id = v.user_id
        WHERE v.deal_id = ?
    """, (deal_id,)).fetchall()
    transfers = db.execute("""
        SELECT t.created_at, du.nome as de_nome, pu.nome as para_nome, au.nome as admin_nome
        FROM customer_transfers t
        LEFT JOIN users du ON du.id = t.de_user_id
        LEFT JOIN users pu ON pu.id = t.para_user_id
        LEFT JOIN users au ON au.id = t.admin_id
        WHERE t.customer_id = ?
    """, (d["customer_id"],)).fetchall()
    eh_admin = g.current_user["role"] == "admin"
    eventos = (
        [{"tipo": "nota", "id": n["id"], "conteudo": n["conteudo"], "etapa_funil": n["etapa_funil"],
          "autor_nome": n["autor_nome"], "autor_id": n["user_id"], "autor_role": n["autor_role"],
          "nota_tipo": n["tipo"], "tem_anexo": bool(n["tem_anexo"]), "data": n["created_at"]}
         for n in notas]
        + [{"tipo": "etapa", "etapa_anterior": e["etapa_anterior"], "etapa_nova": e["etapa_nova"],
            "autor_nome": e["autor_nome"], "data": e["data_transicao"],
            "motivo_perda": e["motivo_perda"], "motivo_perda_detalhe": e["motivo_perda_detalhe"]}
           for e in etapas]
        + [{"tipo": "valor", "valor_anterior": v["valor_anterior"], "valor_novo": v["valor_novo"],
            "autor_nome": v["autor_nome"], "data": v["created_at"]}
           for v in valores]
        + [{"tipo": "transferencia", "data": t["created_at"], "admin_nome": t["admin_nome"],
            "de_nome": t["de_nome"] if eh_admin else None,
            "para_nome": t["para_nome"] if eh_admin else None}
           for t in transfers]
    )
    eventos.sort(key=lambda ev: (ev["data"], 0 if ev["tipo"] == "etapa" else 1))
    return jsonify({"deal": dict(d), "eventos": eventos})


@app.post("/api/deals/<deal_id>/notas")
@login_required
def add_deal_nota(deal_id):
    d, erro = _deal_permitido_ou_erro(deal_id, escrita=True)
    if erro:
        return erro
    body = request.get_json(force=True, silent=True) or {}
    conteudo = (body.get("conteudo") or "").strip()
    tipo = body.get("tipo") or "nota"
    if tipo not in ("nota", "whatsapp"):
        return jsonify({"error": "Tipo de nota inválido."}), 400
    anexo = body.get("anexo") or None

    if not conteudo and not anexo:
        return jsonify({"error": "Escreva o que foi conversado ou anexe um print antes de salvar."}), 400
    # Conversa colada do WhatsApp pode ser longa; anotação comum é mais curta.
    limite = 20000 if tipo == "whatsapp" else 4000
    if len(conteudo) > limite:
        return jsonify({"error": f"O texto é longo demais (máximo de {limite} caracteres para este tipo de nota)."}), 400

    dados_anexo, mime_anexo, nome_anexo = None, None, None
    if anexo:
        mime_anexo = (anexo.get("mime") or "").lower()
        if mime_anexo not in ANEXO_MIMES_PERMITIDOS:
            return jsonify({"error": "O anexo precisa ser uma imagem PNG, JPEG ou WebP."}), 400
        try:
            dados_anexo = base64.b64decode(anexo.get("dados_base64") or "", validate=True)
        except Exception:
            return jsonify({"error": "Não foi possível ler o anexo enviado."}), 400
        if not dados_anexo or len(dados_anexo) > ANEXO_TAMANHO_MAX:
            return jsonify({"error": "O anexo precisa ter no máximo 2 MB."}), 400
        nome_anexo = (anexo.get("nome") or "anexo")[:120]

    db = get_db()
    nid = new_id()
    # A nota fica carimbada com a etapa em que o negócio está AGORA —
    # assim o histórico mostra o que foi conversado em cada fase da venda.
    db.execute("""
        INSERT INTO deal_notes (id, deal_id, user_id, etapa_funil, conteudo, tipo, tem_anexo)
        VALUES (?,?,?,?,?,?,?)
    """, (nid, deal_id, g.current_user["id"], d["etapa_funil"], conteudo, tipo, 1 if dados_anexo else 0))
    if dados_anexo:
        db.execute("INSERT INTO deal_note_anexos (note_id, mime, nome_arquivo, dados) VALUES (?,?,?,?)",
                   (nid, mime_anexo, nome_anexo, dados_anexo))
    audit("create", "deal_notes", nid, {"deal_id": deal_id, "tipo": tipo, "tem_anexo": bool(dados_anexo)})
    db.commit()
    return jsonify({"id": nid}), 201


ANEXO_MIMES_PERMITIDOS = {"image/png", "image/jpeg", "image/webp"}
ANEXO_TAMANHO_MAX = 2 * 1024 * 1024  # 2 MB


@app.get("/api/deals/<deal_id>/notas/<note_id>/anexo")
@login_required
def get_deal_nota_anexo(deal_id, note_id):
    """Serve o print anexado a uma nota, com a MESMA regra de visibilidade
    do histórico (vendedor: só os próprios negócios; admin: todos)."""
    d, erro = _deal_permitido_ou_erro(deal_id)
    if erro:
        return erro
    db = get_db()
    row = db.execute("""
        SELECT a.mime, a.dados FROM deal_note_anexos a
        JOIN deal_notes n ON n.id = a.note_id
        WHERE a.note_id = ? AND n.deal_id = ?
    """, (note_id, deal_id)).fetchone()
    if not row:
        return jsonify({"error": "Anexo não encontrado."}), 404
    return Response(row["dados"], mimetype=row["mime"],
                    headers={"Cache-Control": "private, max-age=300"})


@app.put("/api/deals/<deal_id>")
@login_required
def update_deal(deal_id):
    db = get_db()
    existing = db.execute("SELECT * FROM deals WHERE id = ?", (deal_id,)).fetchone()
    if not existing:
        return jsonify({"error": "Negócio não encontrado."}), 404
    if g.current_user["role"] != "admin" and existing["user_id"] != g.current_user["id"]:
        return jsonify({"error": "Sem permissão para editar este negócio."}), 403
    body = request.get_json(force=True, silent=True) or {}
    titulo = body.get("titulo", existing["titulo"])
    valor = body.get("valor_estimado", existing["valor_estimado"])
    data_prevista = body.get("data_prevista_fechamento", existing["data_prevista_fechamento"])
    produto_qtd = existing["produto_qtd"]
    if existing["produto_id"] and "produto_qtd" in body:
        produto_qtd = _int_or_none(body.get("produto_qtd")) or 1
    db.execute("""
        UPDATE deals SET titulo = ?, valor_estimado = ?, data_prevista_fechamento = ?,
            produto_qtd = ?, updated_at = ?
        WHERE id = ?
    """, (titulo, valor, data_prevista, produto_qtd, now_iso(), deal_id))
    # 💰 mudou o valor? entra no histórico (imutável, como o resto)
    try:
        valor_num = float(valor or 0)
    except (TypeError, ValueError):
        valor_num = 0.0
    valor_antigo = float(existing["valor_estimado"] or 0)
    if valor_num != valor_antigo:
        db.execute("""
            INSERT INTO deal_value_history (id, deal_id, valor_anterior, valor_novo, user_id, created_at)
            VALUES (?,?,?,?,?,?)
        """, (new_id(), deal_id, valor_antigo, valor_num, g.current_user["id"], now_iso()))
    audit("update", "deals", deal_id, body)
    db.commit()
    return jsonify({"ok": True})


# ------------------------------------------------------------------
# Rotas — Assistente de Vendas (motor de regras gratuito, sem API externa)
# ------------------------------------------------------------------
@app.get("/api/assistant/overview")
@login_required
def assistant_overview():
    db = get_db()
    clause, params = scope_filter_clause("user_id")
    deals = db.execute(f"""
        SELECT deals.*, customers.nome as cliente_nome, customers.status_fidelidade, u.nome as vendedor_nome
        FROM deals
        JOIN customers ON customers.id = deals.customer_id
        LEFT JOIN users u ON u.id = deals.user_id
        WHERE deals.status = 'aberto' {clause}
        ORDER BY deals.created_at DESC
    """, params).fetchall()

    resultado = []
    for d in deals:
        customer = db.execute("SELECT * FROM customers WHERE id = ?", (d["customer_id"],)).fetchone()
        tasks = db.execute("SELECT * FROM tasks WHERE deal_id = ?", (d["id"],)).fetchall()
        score, recomendacoes = build_deal_recommendation(d, customer, tasks)
        resultado.append({
            "deal_id": d["id"], "titulo": d["titulo"], "cliente_nome": d["cliente_nome"],
            "vendedor_nome": d["vendedor_nome"],
            "customer_id": d["customer_id"], "etapa_funil": d["etapa_funil"],
            "valor_estimado": d["valor_estimado"], "score": score,
            "top_recomendacao": recomendacoes[0]["texto"] if recomendacoes else None,
            "qtd_recomendacoes": len(recomendacoes),
        })

    resultado.sort(key=lambda r: r["score"], reverse=True)
    prontos = [r for r in resultado if r["score"] >= 70]
    em_risco = [r for r in resultado if r["score"] < 40]

    return jsonify({
        "total_negocios_abertos": len(resultado),
        "prontos_para_fechar": prontos[:5],
        "em_risco": em_risco[:5],
        "todos": resultado,
        "dicas_gerais": DICAS_GERAIS,
    })


@app.get("/api/assistant/deals/<deal_id>")
@login_required
def assistant_deal_detail(deal_id):
    db = get_db()
    d = db.execute("""
        SELECT deals.*, customers.nome as cliente_nome FROM deals
        JOIN customers ON customers.id = deals.customer_id
        WHERE deals.id = ?
    """, (deal_id,)).fetchone()
    if not d:
        return jsonify({"error": "Negócio não encontrado."}), 404
    if g.current_user["role"] != "admin" and d["user_id"] != g.current_user["id"]:
        return jsonify({"error": "Sem permissão para ver este negócio."}), 403
    customer = db.execute("SELECT * FROM customers WHERE id = ?", (d["customer_id"],)).fetchone()
    tasks = db.execute("SELECT * FROM tasks WHERE deal_id = ?", (d["id"],)).fetchall()
    score, recomendacoes = build_deal_recommendation(d, customer, tasks)
    return jsonify({
        "deal_id": d["id"], "titulo": d["titulo"], "cliente_nome": d["cliente_nome"],
        "score": score, "recomendacoes": recomendacoes,
    })


# ------------------------------------------------------------------
# Rotas — Tasks
# ------------------------------------------------------------------
@app.get("/api/tasks")
@login_required
def list_tasks():
    db = get_db()
    clause, params = scope_filter_clause("user_id")
    rows = db.execute(f"""
        SELECT t.*, c.nome as cliente_nome, c.cpf_cnpj as cliente_cpf_cnpj,
               u.nome as vendedor_nome FROM tasks t
        LEFT JOIN customers c ON c.id = t.customer_id
        LEFT JOIN users u ON u.id = t.user_id
        WHERE 1=1 {clause}
        ORDER BY t.executado ASC, t.data_lembrete ASC
    """, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/tasks")
@login_required
def create_task():
    body = request.get_json(force=True, silent=True) or {}
    if not body.get("descricao") or not body.get("data_lembrete"):
        return jsonify({"error": "descricao e data_lembrete são obrigatórios."}), 400
    db = get_db()

    # Se a tarefa está vinculada a um cliente ou negócio, ele precisa pertencer ao usuário
    # (a não ser que seja admin) — evita plantar compromissos na carteira de outro vendedor.
    if body.get("customer_id") and g.current_user["role"] != "admin":
        cliente = db.execute("SELECT * FROM customers WHERE id = ?", (body["customer_id"],)).fetchone()
        if not cliente or cliente["responsavel_id"] != g.current_user["id"]:
            return jsonify({"error": "Você só pode criar compromissos para os seus próprios clientes."}), 403
    if body.get("deal_id") and g.current_user["role"] != "admin":
        deal = db.execute("SELECT * FROM deals WHERE id = ?", (body["deal_id"],)).fetchone()
        if not deal or deal["user_id"] != g.current_user["id"]:
            return jsonify({"error": "Você só pode criar compromissos para os seus próprios negócios."}), 403

    # Vendedor só pode criar compromisso para si mesmo; só admin pode atribuir a outro usuário.
    user_id = body.get("user_id", g.current_user["id"])
    if g.current_user["role"] != "admin":
        user_id = g.current_user["id"]

    tid = new_id()
    tipo_atividade = body.get("tipo_atividade", "outro")
    if tipo_atividade not in ("reuniao", "proposta", "follow_up", "ligacao", "email", "outro"):
        tipo_atividade = "outro"
    db.execute("""
        INSERT INTO tasks (id, deal_id, customer_id, user_id, descricao, tipo_atividade, data_lembrete)
        VALUES (?,?,?,?,?,?,?)
    """, (tid, body.get("deal_id"), body.get("customer_id"), user_id,
          body["descricao"], tipo_atividade, body["data_lembrete"]))
    audit("create", "tasks", tid, body)
    db.commit()
    return jsonify({"id": tid}), 201


@app.post("/api/tasks/<task_id>/complete")
@login_required
def complete_task(task_id):
    db = get_db()
    existing = db.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
    if not existing:
        return jsonify({"error": "Compromisso não encontrado."}), 404
    if g.current_user["role"] != "admin" and existing["user_id"] != g.current_user["id"]:
        return jsonify({"error": "Sem permissão para concluir este compromisso."}), 403
    body = request.get_json(force=True, silent=True) or {}
    nota = (body.get("nota_conclusao") or "").strip() or None
    if nota and len(nota) > 2000:
        return jsonify({"error": "A anotação é longa demais (máximo de 2.000 caracteres)."}), 400
    db.execute("UPDATE tasks SET executado = 1, executado_em = ?, nota_conclusao = ? WHERE id = ?",
               (now_iso(), nota, task_id))
    audit("update", "tasks", task_id, {"executado": True, "com_nota": bool(nota)})
    db.commit()
    return jsonify({"ok": True})


@app.post("/api/tasks/excluir")
@login_required
def delete_completed_tasks():
    """Exclui em lote compromissos JÁ CONCLUÍDOS — limpeza da lista e do
    banco. Vendedor só exclui os próprios; admin exclui qualquer um.
    Compromissos pendentes nunca são excluídos por esta rota (a cláusula
    executado = 1 garante isso mesmo se o id de um pendente for enviado)."""
    body = request.get_json(force=True, silent=True) or {}
    ids = body.get("ids") or []
    if not isinstance(ids, list) or not ids:
        return jsonify({"error": "Informe a lista de compromissos a excluir."}), 400
    if len(ids) > 500:
        return jsonify({"error": "Exclua no máximo 500 compromissos por vez."}), 400
    db = get_db()
    placeholders = ",".join("?" * len(ids))
    params = list(ids)
    filtro_dono = ""
    if g.current_user["role"] != "admin":
        filtro_dono = " AND user_id = ?"
        params.append(g.current_user["id"])
    cur = db.execute(
        f"DELETE FROM tasks WHERE id IN ({placeholders}) AND executado = 1{filtro_dono}",
        params,
    )
    audit("delete", "tasks", "lote", {"solicitadas": len(ids), "excluidas": cur.rowcount})
    db.commit()
    return jsonify({"excluidas": cur.rowcount})


# ------------------------------------------------------------------
# Rotas — Users (admin)
# ------------------------------------------------------------------
@app.get("/api/users")
@login_required
@admin_required
def list_users():
    db = get_db()
    rows = db.execute("SELECT id, nome, email, role, ativo FROM users ORDER BY nome").fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/users")
@login_required
@admin_required
def create_user():
    body = request.get_json(force=True, silent=True) or {}
    for f in ("nome", "email", "senha", "role"):
        if not body.get(f):
            return jsonify({"error": f"Campo '{f}' é obrigatório."}), 400
    if body["role"] not in ("admin", "vendedor"):
        return jsonify({"error": "role inválida."}), 400
    db = get_db()
    uid = new_id()
    try:
        db.execute("""
            INSERT INTO users (id, nome, email, senha_hash, role) VALUES (?,?,?,?,?)
        """, (uid, body["nome"], body["email"].lower(), generate_password_hash(body["senha"]), body["role"]))
    except sqlite3.IntegrityError:
        return jsonify({"error": "Já existe um usuário com esse email."}), 409
    audit("create", "users", uid, {"nome": body["nome"], "role": body["role"]})
    db.commit()
    return jsonify({"id": uid}), 201


@app.delete("/api/users/<user_id>")
@login_required
@admin_required
def deactivate_user(user_id):
    db = get_db()
    target = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not target:
        return jsonify({"error": "Usuário não encontrado."}), 404
    if target["role"] == "admin" and target["ativo"] == 1:
        if active_admin_count(db, exclude_user_id=user_id) < 2:
            return jsonify({"error": "O sistema deve manter no mínimo 2 administradores ativos."}), 409
    db.execute("UPDATE users SET ativo = 0 WHERE id = ?", (user_id,))
    audit("delete", "users", user_id)
    db.commit()
    return jsonify({"ok": True})


@app.put("/api/users/<user_id>")
@login_required
@admin_required
def update_user(user_id):
    db = get_db()
    target = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not target:
        return jsonify({"error": "Usuário não encontrado."}), 404
    body = request.get_json(force=True, silent=True) or {}

    nome = body.get("nome", target["nome"])
    email = (body.get("email", target["email"]) or "").lower()
    role = body.get("role", target["role"])
    ativo_raw = body.get("ativo", target["ativo"])
    ativo = 1 if ativo_raw in (1, "1", True, "true") else 0

    if role not in ("admin", "vendedor"):
        return jsonify({"error": "role inválida."}), 400

    # Se isso tira o usuário da condição de admin ativo, valida a trava dos 2 admins
    deixa_de_ser_admin_ativo = target["role"] == "admin" and target["ativo"] == 1 and (role != "admin" or ativo == 0)
    if deixa_de_ser_admin_ativo and active_admin_count(db, exclude_user_id=user_id) < 2:
        return jsonify({"error": "O sistema deve manter no mínimo 2 administradores ativos."}), 409

    updates = {"nome": nome, "email": email, "role": role, "ativo": ativo}
    if body.get("senha"):
        updates["senha_hash"] = generate_password_hash(body["senha"])

    set_clause = ", ".join(f"{k} = ?" for k in updates)
    try:
        db.execute(f"UPDATE users SET {set_clause}, updated_at = ? WHERE id = ?",
                   (*updates.values(), now_iso(), user_id))
    except sqlite3.IntegrityError:
        return jsonify({"error": "Já existe um usuário com esse email."}), 409

    audit("update", "users", user_id, {k: v for k, v in body.items() if k != "senha"})
    db.commit()
    return jsonify({"ok": True})


# ------------------------------------------------------------------
# Rotas — Tarefas Delegadas (admin atribui, vendedor conduz o fluxo)
# ------------------------------------------------------------------
DELEGATED_STATUS = ("aberta", "em_andamento", "finalizada")


@app.get("/api/delegated-tasks")
@login_required
def list_delegated_tasks():
    db = get_db()
    base_query = """
        SELECT dt.*, u.nome as atribuido_nome, cr.nome as criado_por_nome, c.nome as cliente_nome
        FROM delegated_tasks dt
        JOIN users u ON u.id = dt.atribuido_para
        JOIN users cr ON cr.id = dt.criado_por
        LEFT JOIN customers c ON c.id = dt.customer_id
    """
    if g.current_user["role"] == "admin":
        atribuido = request.args.get("atribuido_para")
        if atribuido:
            rows = db.execute(base_query + " WHERE dt.atribuido_para = ? ORDER BY dt.created_at DESC",
                               (atribuido,)).fetchall()
        else:
            rows = db.execute(base_query + " ORDER BY dt.created_at DESC").fetchall()
    else:
        rows = db.execute(base_query + " WHERE dt.atribuido_para = ? ORDER BY dt.created_at DESC",
                           (g.current_user["id"],)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/delegated-tasks")
@login_required
@admin_required
def create_delegated_task():
    body = request.get_json(force=True, silent=True) or {}
    if not body.get("titulo"):
        return jsonify({"error": "titulo é obrigatório."}), 400
    db = get_db()

    # Delegação em grupo: 'vendedores' (só a equipe de vendas) ou 'todos'
    # (equipe completa, incluindo administradores). Cada pessoa recebe a
    # própria cópia, identificada pelo mesmo grupo_id, e conduz o fluxo
    # de forma independente. 'atribuir_todos' é aceito por compatibilidade
    # com a versão anterior (equivale a 'vendedores').
    equipe = body.get("atribuir_equipe") or ("vendedores" if body.get("atribuir_todos") else None)
    if equipe:
        if equipe not in ("vendedores", "todos"):
            return jsonify({"error": "Grupo inválido: use 'vendedores' ou 'todos'."}), 400
        if equipe == "todos":
            destinatarios = db.execute("SELECT id FROM users WHERE ativo = 1").fetchall()
        else:
            destinatarios = db.execute(
                "SELECT id FROM users WHERE role = 'vendedor' AND ativo = 1"
            ).fetchall()
        if not destinatarios:
            return jsonify({"error": "Não há usuários ativos para atribuir a tarefa."}), 400
        grupo_id = new_id()
        ids_criados = []
        for u in destinatarios:
            tid = new_id()
            db.execute("""
                INSERT INTO delegated_tasks (id, titulo, descricao, criado_por, atribuido_para, customer_id, data_prazo, grupo_id)
                VALUES (?,?,?,?,?,?,?,?)
            """, (tid, body["titulo"], body.get("descricao"), g.current_user["id"], u["id"],
                  body.get("customer_id"), body.get("data_prazo"), grupo_id))
            ids_criados.append(tid)
        audit("create", "delegated_tasks", grupo_id, {**body, "equipe": equipe, "qtd": len(ids_criados)})
        db.commit()
        return jsonify({"grupo_id": grupo_id, "ids": ids_criados, "qtd_atribuidos": len(ids_criados)}), 201

    if not body.get("atribuido_para"):
        return jsonify({"error": "Escolha um responsável ou uma das opções de equipe."}), 400
    destino = db.execute("SELECT id FROM users WHERE id = ? AND ativo = 1",
                         (body["atribuido_para"],)).fetchone()
    if not destino:
        return jsonify({"error": "Responsável inválido ou inativo."}), 400
    tid = new_id()
    db.execute("""
        INSERT INTO delegated_tasks (id, titulo, descricao, criado_por, atribuido_para, customer_id, data_prazo)
        VALUES (?,?,?,?,?,?,?)
    """, (tid, body["titulo"], body.get("descricao"), g.current_user["id"], body["atribuido_para"],
          body.get("customer_id"), body.get("data_prazo")))
    audit("create", "delegated_tasks", tid, body)
    db.commit()
    return jsonify({"id": tid}), 201


@app.put("/api/delegated-tasks/<task_id>")
@login_required
@admin_required
def update_delegated_task(task_id):
    db = get_db()
    t = db.execute("SELECT * FROM delegated_tasks WHERE id = ?", (task_id,)).fetchone()
    if not t:
        return jsonify({"error": "Tarefa não encontrada."}), 404
    body = request.get_json(force=True, silent=True) or {}
    titulo = body.get("titulo", t["titulo"])
    descricao = body.get("descricao", t["descricao"])
    atribuido = body.get("atribuido_para", t["atribuido_para"])
    prazo = body.get("data_prazo", t["data_prazo"])
    db.execute("""
        UPDATE delegated_tasks SET titulo=?, descricao=?, atribuido_para=?, data_prazo=?, updated_at=?
        WHERE id=?
    """, (titulo, descricao, atribuido, prazo, now_iso(), task_id))
    audit("update", "delegated_tasks", task_id, body)
    db.commit()
    return jsonify({"ok": True})


@app.put("/api/delegated-tasks/<task_id>/status")
@login_required
def update_delegated_task_status(task_id):
    db = get_db()
    t = db.execute("SELECT * FROM delegated_tasks WHERE id = ?", (task_id,)).fetchone()
    if not t:
        return jsonify({"error": "Tarefa não encontrada."}), 404
    if g.current_user["role"] != "admin" and t["atribuido_para"] != g.current_user["id"]:
        return jsonify({"error": "Sem permissão para atualizar esta tarefa."}), 403

    body = request.get_json(force=True, silent=True) or {}
    novo_status = body.get("status")
    if novo_status not in DELEGATED_STATUS:
        return jsonify({"error": "Status inválido."}), 400

    descricao_conclusao = body.get("descricao_conclusao", t["descricao_conclusao"])
    if novo_status == "finalizada" and not (descricao_conclusao or "").strip():
        return jsonify({"error": "Descreva brevemente o que foi feito para finalizar a tarefa."}), 400

    finalizada_em = now_iso() if novo_status == "finalizada" else t["finalizada_em"]
    db.execute("""
        UPDATE delegated_tasks SET status = ?, descricao_conclusao = ?, finalizada_em = ?, updated_at = ?
        WHERE id = ?
    """, (novo_status, descricao_conclusao, finalizada_em, now_iso(), task_id))
    audit("update", "delegated_tasks", task_id, {"status": novo_status})
    db.commit()
    return jsonify({"ok": True})


@app.get("/api/lembretes")
@login_required
def list_lembretes():
    """Lembretes pessoais de tarefas para o Feed de Alertas. São SEMPRE do
    próprio usuário logado — por isso as regras de visibilidade se resolvem
    sozinhas: vendedor vê os próprios compromissos e as tarefas que o admin
    delegou a ele; admin vê as suas (inclusive as trocadas entre admins);
    ninguém vê lembrete de outra pessoa.
    - Compromissos: os que vencem nas próximas 24h. Os já vencidos NÃO
      entram aqui — viram o alerta 'compromisso esquecido' pelo motor de
      alertas, sem duplicar.
    - Delegadas: pendentes com prazo até amanhã, incluindo as vencidas
      (que o motor de alertas não cobre)."""
    db = get_db()
    uid = g.current_user["id"]
    agora = now_iso()
    limite_24h = (datetime.now(timezone.utc) + timedelta(hours=24)).strftime("%Y-%m-%d %H:%M:%S")
    compromissos = db.execute("""
        SELECT t.*, c.nome as cliente_nome FROM tasks t
        LEFT JOIN customers c ON c.id = t.customer_id
        WHERE t.user_id = ? AND t.executado = 0
          AND t.data_lembrete > ? AND t.data_lembrete <= ?
        ORDER BY t.data_lembrete
    """, (uid, agora, limite_24h)).fetchall()
    ate_amanha = (datetime.now(timezone.utc) + timedelta(days=1)).strftime("%Y-%m-%d")
    delegadas = db.execute("""
        SELECT dt.*, cr.nome as criado_por_nome, c.nome as cliente_nome
        FROM delegated_tasks dt
        JOIN users cr ON cr.id = dt.criado_por
        LEFT JOIN customers c ON c.id = dt.customer_id
        WHERE dt.atribuido_para = ? AND dt.status != 'finalizada'
          AND dt.data_prazo IS NOT NULL AND dt.data_prazo <= ?
        ORDER BY dt.data_prazo
    """, (uid, ate_amanha)).fetchall()
    # 🎯 Metas ativas do usuário ainda NÃO atingidas (bloco 📌 do feed)
    likes_eq, params_eq = _clientes_sem_oferta_sql("customers")
    sem_oferta = db.execute(f"""
        SELECT COUNT(*) c FROM customers
        WHERE ativo = 1 AND responsavel_id = ? AND ({likes_eq})
          AND NOT EXISTS (SELECT 1 FROM deals d WHERE d.customer_id = customers.id AND d.categoria = 'software')
    """, [uid] + params_eq).fetchone()["c"]
    metas_pendentes = []
    for m in _metas_do_usuario(db, uid, g.current_user["role"]):
        prog = _meta_progresso_map(db, m)
        parts_ids = [p["id"] for p in _meta_participantes(db, m)]
        meu = prog.get(uid, 0)
        total = sum(q for u, q in prog.items() if u in parts_ids)
        efetivo = total if m["apuracao"] == "coletiva" else meu
        if efetivo >= m["quantidade"]:
            continue
        eh_software = m["tipo"] == "vendas" and (m["alvo_vendas"] or "") in (
            "software_qualquer", "revele_momentos", "revele_momentos_frontier")
        metas_pendentes.append({
            "id": m["id"], "titulo": m["titulo"], "quantidade": m["quantidade"],
            "progresso": efetivo, "apuracao": m["apuracao"], "tipo": m["tipo"],
            "periodo_tipo": m["periodo_tipo"], "data_fim": m["data_fim"],
            "clientes_sem_oferta": sem_oferta if eh_software else None,
        })
    # 🙋 liberações de preço: admin vê os pendentes; vendedor vê as decisões
    if g.current_user["role"] == "admin":
        libs = db.execute("""
            SELECT l.id, l.preco_pedido, l.status, p.nome as produto_nome, p.preco_limite,
                   u.nome as vendedor_nome, c.nome as cliente_nome
            FROM liberacoes_preco l
            JOIN produtos p ON p.id = l.produto_id
            JOIN users u ON u.id = l.user_id
            JOIN deals d ON d.id = l.deal_id
            JOIN customers c ON c.id = d.customer_id
            WHERE l.status = 'pendente' ORDER BY l.created_at
        """).fetchall()
    else:
        libs = db.execute("""
            SELECT l.id, l.preco_autorizado as preco_pedido, l.status, p.nome as produto_nome,
                   p.preco_limite, u.nome as vendedor_nome, c.nome as cliente_nome
            FROM liberacoes_preco l
            JOIN produtos p ON p.id = l.produto_id
            JOIN users u ON u.id = l.user_id
            JOIN deals d ON d.id = l.deal_id
            JOIN customers c ON c.id = d.customer_id
            WHERE l.user_id = ? AND l.status IN ('aprovada', 'negada')
              AND l.decidido_em >= ?
            ORDER BY l.decidido_em DESC
        """, (g.current_user["id"],
              (datetime.now(timezone.utc) - timedelta(days=3)).strftime("%Y-%m-%d %H:%M:%S"))).fetchall()
    return jsonify({
        "hoje": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "compromissos": [dict(r) for r in compromissos],
        "delegadas": [dict(r) for r in delegadas],
        "metas": metas_pendentes,
        "liberacoes": [dict(r) for r in libs],
    })


# ------------------------------------------------------------------
# 📦 CATÁLOGO DE PRODUTOS — o elo estruturado entre a venda e a meta.
# Admin cadastra/desativa; produtos nunca são apagados (negócios e metas
# antigos continuam apontando para eles). Todos leem o catálogo, pois o
# vendedor o usa no "Tipo de negócio".
# ------------------------------------------------------------------
@app.get("/api/produtos")
@login_required
def list_produtos():
    db = get_db()
    rows = db.execute("""
        SELECT p.*, u.nome as criado_por_nome FROM produtos p
        LEFT JOIN users u ON u.id = p.criado_por
        ORDER BY p.ativo DESC, p.nome
    """).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/produtos")
@login_required
def create_produto():
    if g.current_user["role"] != "admin":
        return jsonify({"error": "Apenas administradores cadastram produtos."}), 403
    body = request.get_json(force=True, silent=True) or {}
    nome = (body.get("nome") or "").strip()
    if not nome:
        return jsonify({"error": "Informe o nome do produto."}), 400
    if len(nome) > 120:
        return jsonify({"error": "Nome do produto longo demais (máximo 120 caracteres)."}), 400
    db = get_db()
    duplicado = db.execute("SELECT id FROM produtos WHERE lower(nome) = lower(?)", (nome,)).fetchone()
    if duplicado:
        return jsonify({"error": "Já existe um produto com esse nome no catálogo."}), 400
    pid = new_id()
    db.execute("INSERT INTO produtos (id, nome, criado_por, created_at) VALUES (?,?,?,?)",
               (pid, nome, g.current_user["id"], now_iso()))
    audit("create", "produtos", pid, {"nome": nome})
    db.commit()
    return jsonify({"id": pid}), 201


@app.post("/api/produtos/<produto_id>/toggle")
@login_required
def toggle_produto(produto_id):
    """Ativa/desativa um produto do catálogo. Desativado some das opções de
    novo negócio e de nova meta, mas negócios e metas existentes seguem
    intactos (por isso não existe exclusão)."""
    if g.current_user["role"] != "admin":
        return jsonify({"error": "Apenas administradores alteram o catálogo."}), 403
    db = get_db()
    p = db.execute("SELECT * FROM produtos WHERE id = ?", (produto_id,)).fetchone()
    if not p:
        return jsonify({"error": "Produto não encontrado."}), 404
    novo = 0 if p["ativo"] else 1
    db.execute("UPDATE produtos SET ativo = ? WHERE id = ?", (novo, produto_id))
    audit("update", "produtos", produto_id, {"ativo": bool(novo)})
    db.commit()
    return jsonify({"ok": True, "ativo": bool(novo)})


def _num(v):
    """Converte célula da planilha em número (ou None)."""
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace("R$", "").replace(".", "").replace(",", ".")) \
            if isinstance(v, str) and "," in str(v) else float(v)
    except (TypeError, ValueError):
        return None


@app.post("/api/base-comercial/importar")
@login_required
def importar_base_comercial():
    """📥 Importa a planilha oficial (.xlsx) da Base Comercial (admin):
    - Produtos: upsert por nome (atualiza preços/status; cria os novos);
      status contendo "não ofertar" zera a flag ofertavel.
    - Frete por UF e Condições de pagamento: recarga completa (a planilha
      é a fonte canônica dessas regras).
    Nunca apaga produtos: itens fora da planilha permanecem como estão."""
    if g.current_user["role"] != "admin":
        return jsonify({"error": "Apenas administradores importam a base comercial."}), 403
    arquivo = request.files.get("arquivo")
    if not arquivo or not arquivo.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Envie o arquivo .xlsx da Base Comercial."}), 400
    try:
        import openpyxl
        wb = openpyxl.load_workbook(arquivo, data_only=True)
    except Exception:
        return jsonify({"error": "Não consegui ler o arquivo — confirme que é a planilha .xlsx original."}), 400

    db = get_db()
    agora = now_iso()
    resumo = {"produtos_novos": 0, "produtos_atualizados": 0, "fretes": 0, "condicoes": 0, "notas": 0}

    # ---------------- Produtos ----------------
    if "Produtos" in wb.sheetnames:
        ws = wb["Produtos"]
        cabecalho_visto = False
        for row in ws.iter_rows(values_only=True):
            if not cabecalho_visto:
                if row and str(row[0]).strip() == "ID" and "Categoria" in [str(c) for c in row if c]:
                    cabecalho_visto = True
                continue
            if not row or row[0] is None or not row[3]:
                continue
            seq = _int_or_none(row[0])
            categoria = (str(row[1]).strip() if row[1] else None)
            equipamento = (str(row[2]).strip() if row[2] else None)
            nome = str(row[3]).strip()
            embalagem = (str(row[5]).strip() if row[5] else None)
            preco_tabela = _num(row[6])
            preco_limite = _num(row[7])
            desconto_max = _num(row[8])
            status = (str(row[9]).strip() if row[9] else "OK")
            ofertavel = 0 if "não ofertar" in status.lower() else 1
            existente = db.execute("SELECT id FROM produtos WHERE lower(nome) = lower(?)", (nome,)).fetchone()
            if existente:
                db.execute("""
                    UPDATE produtos SET seq = ?, categoria = ?, equipamento = ?, embalagem = ?,
                        preco_tabela = ?, preco_limite = ?, desconto_max = ?, status_comercial = ?,
                        ofertavel = ?
                    WHERE id = ?
                """, (seq, categoria, equipamento, embalagem, preco_tabela, preco_limite,
                      desconto_max, status, ofertavel, existente["id"]))
                resumo["produtos_atualizados"] += 1
            else:
                db.execute("""
                    INSERT INTO produtos (id, nome, ativo, criado_por, created_at, seq, categoria,
                        equipamento, embalagem, preco_tabela, preco_limite, desconto_max,
                        status_comercial, ofertavel)
                    VALUES (?,?,1,?,?,?,?,?,?,?,?,?,?,?)
                """, (new_id(), nome, g.current_user["id"], agora, seq, categoria, equipamento,
                      embalagem, preco_tabela, preco_limite, desconto_max, status, ofertavel))
                resumo["produtos_novos"] += 1

    # ---------------- Frete por UF ----------------
    if "Frete Grátis" in wb.sheetnames:
        ws = wb["Frete Grátis"]
        db.execute("DELETE FROM frete_uf")
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 4 or not row[1]:
                continue
            uf = str(row[1]).strip().upper()
            if len(uf) != 2 or uf == "UF":
                continue
            minimo = _num(row[3])
            if minimo is None:
                continue
            db.execute("""
                INSERT OR REPLACE INTO frete_uf (uf, regiao, estado, minimo, atualizado_em)
                VALUES (?,?,?,?,?)
            """, (uf, str(row[0]).strip() if row[0] else None,
                  str(row[2]).strip() if row[2] else None, minimo, agora))
            resumo["fretes"] += 1

    # ---------------- Condições de pagamento ----------------
    if "Pagamentos" in wb.sheetnames:
        ws = wb["Pagamentos"]
        db.execute("DELETE FROM condicoes_pagamento WHERE simples = 0")
        cabecalho_visto = False
        ordem = 0
        for row in ws.iter_rows(values_only=True):
            if not cabecalho_visto:
                if row and str(row[0]).strip() == "Perfil do cliente":
                    cabecalho_visto = True
                continue
            if not row or not row[0]:
                continue
            perfil = str(row[0]).strip()
            forma = str(row[1]).strip() if row[1] else ""
            condicao = str(row[2]).strip() if row[2] else ""
            regra = str(row[3]).strip() if row[3] else ""
            eh_nota = 1 if (not condicao or perfil.lower() in ("importante", "validade", "regra confirmada")) else 0
            ordem += 1
            ordem_gravar = ordem + 10  # as 3 simplificadas (1-3) vêm antes
            m_pct = re.search(r"acr[eé]scimo de\s*(\d+(?:[.,]\d+)?)\s*%", condicao, re.IGNORECASE)
            acrescimo_pct = float(m_pct.group(1).replace(",", ".")) if m_pct else 0.0
            db.execute("""
                INSERT INTO condicoes_pagamento (id, perfil, forma, condicao, regra, eh_nota, ordem, acrescimo_pct, atualizado_em)
                VALUES (?,?,?,?,?,?,?,?,?)
            """, (new_id(), perfil, forma, condicao, regra, eh_nota, ordem_gravar, acrescimo_pct, agora))
            resumo["notas" if eh_nota else "condicoes"] += 1

    audit("update", "base_comercial", "importacao", resumo)
    db.commit()
    return jsonify({"ok": True, **resumo})


@app.get("/api/base-comercial")
@login_required
def consultar_base_comercial():
    """Consulta de fretes por UF e condições de pagamento (toda a equipe —
    são regras operacionais que o vendedor precisa saber de cor)."""
    db = get_db()
    fretes = [dict(r) for r in db.execute("SELECT * FROM frete_uf ORDER BY minimo, uf").fetchall()]
    condicoes = [dict(r) for r in db.execute(
        "SELECT * FROM condicoes_pagamento ORDER BY eh_nota, ordem").fetchall()]
    return jsonify({"fretes": fretes, "condicoes": condicoes})


@app.put("/api/produtos/<produto_id>")
@login_required
def update_produto(produto_id):
    """Renomeia um produto do catálogo (admin). Como negócios e metas
    referenciam o produto por id, o nome novo reflete em tudo na hora."""
    if g.current_user["role"] != "admin":
        return jsonify({"error": "Apenas administradores alteram o catálogo."}), 403
    db = get_db()
    p = db.execute("SELECT * FROM produtos WHERE id = ?", (produto_id,)).fetchone()
    if not p:
        return jsonify({"error": "Produto não encontrado."}), 404
    body = request.get_json(force=True, silent=True) or {}
    nome = (body.get("nome") or "").strip()
    if not nome:
        return jsonify({"error": "Informe o novo nome do produto."}), 400
    if len(nome) > 120:
        return jsonify({"error": "Nome do produto longo demais (máximo 120 caracteres)."}), 400
    duplicado = db.execute("SELECT id FROM produtos WHERE lower(nome) = lower(?) AND id != ?",
                           (nome, produto_id)).fetchone()
    if duplicado:
        return jsonify({"error": "Já existe outro produto com esse nome no catálogo."}), 400
    db.execute("UPDATE produtos SET nome = ? WHERE id = ?", (nome, produto_id))
    audit("update", "produtos", produto_id, {"nome": nome})
    db.commit()
    return jsonify({"ok": True})


@app.delete("/api/produtos/<produto_id>")
@login_required
def delete_produto(produto_id):
    """Exclui um produto do catálogo (admin) — SOMENTE se nunca foi usado.
    Produto referenciado por negócios ou metas não pode ser apagado, para
    não corromper históricos e relatórios; o caminho para esses é Desativar."""
    if g.current_user["role"] != "admin":
        return jsonify({"error": "Apenas administradores alteram o catálogo."}), 403
    db = get_db()
    p = db.execute("SELECT * FROM produtos WHERE id = ?", (produto_id,)).fetchone()
    if not p:
        return jsonify({"error": "Produto não encontrado."}), 404
    em_negocios = db.execute("SELECT COUNT(*) n FROM deals WHERE produto_id = ?", (produto_id,)).fetchone()["n"]
    em_metas = db.execute("SELECT COUNT(*) n FROM metas WHERE alvo_produto_id = ?", (produto_id,)).fetchone()["n"]
    if em_negocios or em_metas:
        return jsonify({"error": f"Este produto já foi usado em {em_negocios} negócio(s) e {em_metas} meta(s) — para preservar o histórico, ele não pode ser excluído. Use \"desativar\": ele some das opções sem afetar nada do passado."}), 400
    db.execute("DELETE FROM produtos WHERE id = ?", (produto_id,))
    audit("delete", "produtos", produto_id, {"nome": p["nome"]})
    db.commit()
    return jsonify({"ok": True})


# ------------------------------------------------------------------
# 🎯 METAS configuráveis
# - tipo 'vendas': o progresso conta SOZINHO pelos negócios GANHOS na janela
#   de apuração (mês corrente / período definido / desde sempre).
# - tipo 'manual': o usuário registra o progresso (+1 com nota) — histórico.
# - apuração 'individual' (cada participante deve atingir o alvo) ou
#   'coletiva' (a soma do grupo deve atingir o alvo).
# RBAC: admin cria/exclui e vê tudo; vendedor vê apenas as metas dele.
# ------------------------------------------------------------------
ALVOS_VENDA_META = ("qualquer", "padrao", "software_qualquer",
                    "revele_momentos", "revele_momentos_frontier", "produto")


def _meta_situacao(m):
    if m["status"] == "excluida":
        return "excluida"
    hoje = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if m["periodo_tipo"] == "periodo" and m["data_fim"] and m["data_fim"] < hoje:
        return "encerrada"
    return "ativa"


def _meta_janela(m):
    """(inicio, fim) da janela de apuração, comparáveis com timestamps."""
    agora = datetime.now(timezone.utc)
    if m["periodo_tipo"] == "mensal":
        return agora.strftime("%Y-%m-01 00:00:00"), agora.strftime("%Y-%m-%d 23:59:59")
    if m["periodo_tipo"] == "periodo":
        return ((m["data_inicio"] or "1970-01-01") + " 00:00:00",
                (m["data_fim"] or "9999-12-31") + " 23:59:59")
    return "1970-01-01 00:00:00", "9999-12-31 23:59:59"


def _meta_ids_selecionados(m):
    try:
        ids = json.loads(m["escopo_users"] or "[]")
        return ids if isinstance(ids, list) else []
    except Exception:
        return []


def _meta_participantes(db, m):
    if m["escopo"] == "selecionados":
        ids = _meta_ids_selecionados(m)
        if not ids:
            return []
        return db.execute(
            f"SELECT id, nome, role FROM users WHERE id IN ({','.join('?' * len(ids))}) ORDER BY nome",
            ids).fetchall()
    if m["escopo"] == "individual":
        return db.execute("SELECT id, nome, role FROM users WHERE id = ?",
                          (m["escopo_user_id"],)).fetchall()
    if m["escopo"] == "vendedores":
        return db.execute("SELECT id, nome, role FROM users WHERE ativo = 1 AND role = 'vendedor' ORDER BY nome").fetchall()
    return db.execute("SELECT id, nome, role FROM users WHERE ativo = 1 ORDER BY nome").fetchall()


def _meta_filtro_vendas(m):
    alvo = m["alvo_vendas"] or "qualquer"
    if alvo == "produto":
        # só o produto específico do catálogo pontua nesta meta
        return " AND produto_id = ?", [m["alvo_produto_id"]]
    if alvo == "padrao":
        return " AND categoria = 'padrao'", []
    if alvo == "software_qualquer":
        return " AND categoria = 'software'", []
    if alvo in ("revele_momentos", "revele_momentos_frontier"):
        return " AND categoria = 'software' AND produto_software = ?", [alvo]
    return "", []


def _meta_progresso_map(db, m):
    """{user_id: quantidade apurada} dentro da janela."""
    ini, fim = _meta_janela(m)
    if m["tipo"] == "vendas":
        # conta UNIDADES vendidas. Fonte da verdade: os ITENS do orçamento
        # (soma das quantidades). Negócio antigo sem itens usa o produto
        # principal (produto_id/produto_qtd); sem nada, vale 1 unidade.
        if (m["alvo_vendas"] or "qualquer") == "produto":
            pid = m["alvo_produto_id"]
            rows = db.execute("""
                SELECT d.user_id, COALESCE(SUM(
                    CASE WHEN EXISTS (SELECT 1 FROM deal_itens i WHERE i.deal_id = d.id)
                         THEN COALESCE((SELECT SUM(i.qtd) FROM deal_itens i
                                        WHERE i.deal_id = d.id AND i.produto_id = ?), 0)
                         ELSE CASE WHEN d.produto_id = ? THEN COALESCE(d.produto_qtd, 1) ELSE 0 END
                    END), 0) qtd
                FROM deals d
                WHERE d.status = 'ganho' AND d.etapa_atualizada_em BETWEEN ? AND ?
                GROUP BY d.user_id
            """, (pid, pid, ini, fim)).fetchall()
        else:
            extra, extra_params = _meta_filtro_vendas(m)
            rows = db.execute(f"""
                SELECT d.user_id, COALESCE(SUM(
                    CASE WHEN EXISTS (SELECT 1 FROM deal_itens i WHERE i.deal_id = d.id)
                         THEN (SELECT SUM(i.qtd) FROM deal_itens i WHERE i.deal_id = d.id)
                         ELSE COALESCE(d.produto_qtd, 1) END), 0) qtd
                FROM deals d
                WHERE d.status = 'ganho' AND d.etapa_atualizada_em BETWEEN ? AND ?{extra}
                GROUP BY d.user_id
            """, [ini, fim] + extra_params).fetchall()
    else:
        rows = db.execute("""
            SELECT user_id, COALESCE(SUM(quantidade), 0) qtd FROM meta_progresso
            WHERE meta_id = ? AND created_at BETWEEN ? AND ?
            GROUP BY user_id
        """, (m["id"], ini, fim)).fetchall()
    return {r["user_id"]: r["qtd"] for r in rows}


def _metas_do_usuario(db, uid, role):
    """Metas ATIVAS (status + período vigente) aplicáveis ao usuário."""
    hoje = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    out = []
    for m in db.execute("SELECT * FROM metas WHERE status = 'ativa'").fetchall():
        if m["periodo_tipo"] == "periodo":
            if m["data_inicio"] and hoje < m["data_inicio"]:
                continue
            if m["data_fim"] and hoje > m["data_fim"]:
                continue
        if m["escopo"] == "individual" and m["escopo_user_id"] != uid:
            continue
        if m["escopo"] == "selecionados" and uid not in _meta_ids_selecionados(m):
            continue
        if m["escopo"] == "vendedores" and role != "vendedor":
            continue
        out.append(m)
    return out


@app.get("/api/metas")
@login_required
def list_metas():
    db = get_db()
    uid = g.current_user["id"]
    if g.current_user["role"] == "admin":
        metas = db.execute("""
            SELECT m.*, cu.nome as criado_por_nome, eu.nome as escopo_user_nome,
                   ex.nome as excluida_por_nome, ed.nome as editada_por_nome,
                   pr.nome as alvo_produto_nome
            FROM metas m
            LEFT JOIN users cu ON cu.id = m.criado_por
            LEFT JOIN users eu ON eu.id = m.escopo_user_id
            LEFT JOIN users ex ON ex.id = m.excluida_por
            LEFT JOIN users ed ON ed.id = m.editada_por
            LEFT JOIN produtos pr ON pr.id = m.alvo_produto_id
            ORDER BY m.created_at DESC
        """).fetchall()
        saida = []
        for m in metas:
            prog = _meta_progresso_map(db, m)
            parts = _meta_participantes(db, m)
            participantes = [{"user_id": p["id"], "nome": p["nome"], "role": p["role"],
                              "progresso": prog.get(p["id"], 0),
                              "atingida": prog.get(p["id"], 0) >= m["quantidade"]}
                             for p in parts]
            total = sum(p["progresso"] for p in participantes)
            registros = []
            if m["tipo"] == "manual":
                registros = [dict(r) for r in db.execute("""
                    SELECT mp.*, u.nome as autor_nome FROM meta_progresso mp
                    LEFT JOIN users u ON u.id = mp.user_id
                    WHERE mp.meta_id = ? ORDER BY mp.created_at DESC LIMIT 100
                """, (m["id"],)).fetchall()]
            saida.append({**dict(m), "situacao": _meta_situacao(m),
                          "participantes": participantes, "progresso_total": total,
                          "atingida_coletiva": total >= m["quantidade"],
                          "registros": registros})
        return jsonify({"admin": True, "metas": saida})

    saida = []
    for m in _metas_do_usuario(db, uid, g.current_user["role"]):
        prog = _meta_progresso_map(db, m)
        parts_ids = [p["id"] for p in _meta_participantes(db, m)]
        meu = prog.get(uid, 0)
        total = sum(q for u, q in prog.items() if u in parts_ids)
        registros = []
        if m["tipo"] == "manual":
            registros = [dict(r) for r in db.execute("""
                SELECT mp.* FROM meta_progresso mp
                WHERE mp.meta_id = ? AND mp.user_id = ?
                ORDER BY mp.created_at DESC LIMIT 50
            """, (m["id"], uid)).fetchall()]
        nome_prod = None
        if m["alvo_vendas"] == "produto" and m["alvo_produto_id"]:
            pr = db.execute("SELECT nome FROM produtos WHERE id = ?", (m["alvo_produto_id"],)).fetchone()
            nome_prod = pr["nome"] if pr else None
        saida.append({"id": m["id"], "titulo": m["titulo"], "descricao": m["descricao"],
                      "tipo": m["tipo"], "alvo_vendas": m["alvo_vendas"],
                      "alvo_produto_nome": nome_prod,
                      "quantidade": m["quantidade"], "apuracao": m["apuracao"],
                      "periodo_tipo": m["periodo_tipo"], "data_inicio": m["data_inicio"],
                      "data_fim": m["data_fim"], "meu_progresso": meu,
                      "progresso_total": total,
                      "atingida": (total if m["apuracao"] == "coletiva" else meu) >= m["quantidade"],
                      "registros": registros})
    return jsonify({"admin": False, "metas": saida})


def _validar_meta_body(body, db):
    """Valida e normaliza o payload de meta (criação E edição).
    Retorna (campos_dict, None) quando ok, ou (None, (resposta, status))."""
    titulo = (body.get("titulo") or "").strip()
    if not titulo:
        return None, (jsonify({"error": "Dê um título à meta."}), 400)
    tipo = body.get("tipo") or "vendas"
    if tipo not in ("vendas", "manual"):
        return None, (jsonify({"error": "Tipo de meta inválido."}), 400)
    alvo_vendas = None
    alvo_produto_id = None
    if tipo == "vendas":
        alvo_vendas = body.get("alvo_vendas") or "qualquer"
        if alvo_vendas not in ALVOS_VENDA_META:
            return None, (jsonify({"error": "Escolha o que conta como venda para esta meta."}), 400)
        if alvo_vendas == "produto":
            alvo_produto_id = body.get("alvo_produto_id")
            p = db.execute("SELECT id FROM produtos WHERE id = ? AND ativo = 1",
                           (alvo_produto_id,)).fetchone()
            if not p:
                return None, (jsonify({"error": "Escolha um produto ativo do catálogo para esta meta."}), 400)
    quantidade = _int_or_none(body.get("quantidade"))
    if not quantidade:
        return None, (jsonify({"error": "Informe a quantidade alvo (número maior que zero)."}), 400)
    apuracao = body.get("apuracao") or "individual"
    if apuracao not in ("individual", "coletiva"):
        return None, (jsonify({"error": "Apuração inválida."}), 400)
    escopo = body.get("escopo") or "todos"
    if escopo not in ("todos", "vendedores", "individual", "selecionados"):
        return None, (jsonify({"error": "Escopo inválido."}), 400)
    escopo_user_id = None
    escopo_users = None
    if escopo == "individual":
        escopo_user_id = body.get("escopo_user_id")
        u = db.execute("SELECT id FROM users WHERE id = ? AND ativo = 1", (escopo_user_id,)).fetchone()
        if not u:
            return None, (jsonify({"error": "Para meta individual, escolha o usuário no campo correspondente."}), 400)
    if escopo == "selecionados":
        ids = body.get("escopo_users")
        if not isinstance(ids, list) or not ids:
            return None, (jsonify({"error": "Marque ao menos um usuário para a meta de usuários específicos."}), 400)
        ids = list(dict.fromkeys(ids))  # remove duplicados preservando a ordem
        validos = db.execute(
            f"SELECT COUNT(*) c FROM users WHERE ativo = 1 AND id IN ({','.join('?' * len(ids))})",
            ids).fetchone()["c"]
        if validos != len(ids):
            return None, (jsonify({"error": "Um dos usuários marcados é inválido ou está inativo."}), 400)
        escopo_users = json.dumps(ids)
    periodo_tipo = body.get("periodo_tipo") or "mensal"
    if periodo_tipo not in ("mensal", "periodo", "sem_fim"):
        return None, (jsonify({"error": "Período inválido."}), 400)
    data_inicio = data_fim = None
    if periodo_tipo == "periodo":
        data_inicio = (body.get("data_inicio") or "").strip() or None
        data_fim = (body.get("data_fim") or "").strip() or None
        if not data_inicio or not data_fim or data_fim < data_inicio:
            return None, (jsonify({"error": "Período definido exige data de início e de fim (fim após o início)."}), 400)
    return {
        "titulo": titulo, "descricao": (body.get("descricao") or "").strip() or None,
        "tipo": tipo, "alvo_vendas": alvo_vendas, "alvo_produto_id": alvo_produto_id,
        "quantidade": quantidade, "apuracao": apuracao, "escopo": escopo,
        "escopo_user_id": escopo_user_id, "escopo_users": escopo_users,
        "periodo_tipo": periodo_tipo, "data_inicio": data_inicio, "data_fim": data_fim,
    }, None


@app.post("/api/metas")
@login_required
def create_meta():
    if g.current_user["role"] != "admin":
        return jsonify({"error": "Apenas administradores criam metas."}), 403
    body = request.get_json(force=True, silent=True) or {}
    db = get_db()
    campos, erro = _validar_meta_body(body, db)
    if erro:
        return erro
    titulo = campos["titulo"]
    tipo = campos["tipo"]
    alvo_vendas = campos["alvo_vendas"]
    alvo_produto_id = campos["alvo_produto_id"]
    quantidade = campos["quantidade"]
    apuracao = campos["apuracao"]
    escopo = campos["escopo"]
    escopo_user_id = campos["escopo_user_id"]
    periodo_tipo = campos["periodo_tipo"]
    data_inicio = campos["data_inicio"]
    data_fim = campos["data_fim"]
    body = {**body, "descricao": campos["descricao"]}
    mid = new_id()
    db.execute("""
        INSERT INTO metas (id, titulo, descricao, tipo, alvo_vendas, alvo_produto_id, quantidade,
                           apuracao, escopo, escopo_user_id, escopo_users, periodo_tipo,
                           data_inicio, data_fim, criado_por, created_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (mid, titulo, campos["descricao"], tipo, alvo_vendas,
          alvo_produto_id, quantidade, apuracao, escopo, escopo_user_id, campos["escopo_users"],
          periodo_tipo, data_inicio, data_fim, g.current_user["id"], now_iso()))
    audit("create", "metas", mid, {"titulo": titulo, "escopo": escopo, "apuracao": apuracao,
                                   "periodo": periodo_tipo})
    db.commit()
    return jsonify({"id": mid}), 201


@app.put("/api/metas/<meta_id>")
@login_required
def update_meta(meta_id):
    """Edita uma meta existente (admin). Metas excluídas não são editáveis.
    O progresso é recalculado ao vivo com as novas regras; estender a data
    fim de uma meta encerrada a reativa automaticamente."""
    if g.current_user["role"] != "admin":
        return jsonify({"error": "Apenas administradores editam metas."}), 403
    db = get_db()
    m = db.execute("SELECT * FROM metas WHERE id = ?", (meta_id,)).fetchone()
    if not m:
        return jsonify({"error": "Meta não encontrada."}), 404
    if m["status"] == "excluida":
        return jsonify({"error": "Meta excluída não pode ser editada — crie uma nova."}), 400
    body = request.get_json(force=True, silent=True) or {}
    campos, erro = _validar_meta_body(body, db)
    if erro:
        return erro
    db.execute("""
        UPDATE metas SET titulo = ?, descricao = ?, tipo = ?, alvo_vendas = ?, alvo_produto_id = ?,
            quantidade = ?, apuracao = ?, escopo = ?, escopo_user_id = ?, escopo_users = ?,
            periodo_tipo = ?, data_inicio = ?, data_fim = ?, editada_por = ?, editada_em = ?
        WHERE id = ?
    """, (campos["titulo"], campos["descricao"], campos["tipo"], campos["alvo_vendas"],
          campos["alvo_produto_id"], campos["quantidade"], campos["apuracao"], campos["escopo"],
          campos["escopo_user_id"], campos["escopo_users"], campos["periodo_tipo"],
          campos["data_inicio"], campos["data_fim"], g.current_user["id"], now_iso(), meta_id))
    audit("update", "metas", meta_id, {"titulo": campos["titulo"], "escopo": campos["escopo"],
                                       "periodo": campos["periodo_tipo"]})
    db.commit()
    return jsonify({"ok": True})


@app.delete("/api/metas/<meta_id>")
@login_required
def delete_meta(meta_id):
    if g.current_user["role"] != "admin":
        return jsonify({"error": "Apenas administradores excluem metas."}), 403
    db = get_db()
    m = db.execute("SELECT * FROM metas WHERE id = ?", (meta_id,)).fetchone()
    if not m:
        return jsonify({"error": "Meta não encontrada."}), 404
    if m["status"] == "excluida":
        return jsonify({"error": "Esta meta já foi excluída."}), 400
    # exclusão LÓGICA: sai dos alertas/cobrança, permanece na lista histórica
    db.execute("UPDATE metas SET status = 'excluida', excluida_por = ?, excluida_em = ? WHERE id = ?",
               (g.current_user["id"], now_iso(), meta_id))
    audit("delete", "metas", meta_id, {"titulo": m["titulo"]})
    db.commit()
    return jsonify({"ok": True})


@app.post("/api/metas/<meta_id>/progresso")
@login_required
def add_meta_progresso(meta_id):
    db = get_db()
    m = db.execute("SELECT * FROM metas WHERE id = ?", (meta_id,)).fetchone()
    if not m:
        return jsonify({"error": "Meta não encontrada."}), 404
    if m["tipo"] != "manual":
        return jsonify({"error": "Esta meta é automática: o progresso conta sozinho a cada venda ganha."}), 400
    if _meta_situacao(m) != "ativa":
        return jsonify({"error": "Esta meta não está mais ativa."}), 400
    if g.current_user["id"] not in [p["id"] for p in _meta_participantes(db, m)]:
        return jsonify({"error": "Esta meta não se aplica a você."}), 403
    body = request.get_json(force=True, silent=True) or {}
    nota = (body.get("nota") or "").strip() or None
    if nota and len(nota) > 1000:
        return jsonify({"error": "A nota é longa demais (máximo de 1.000 caracteres)."}), 400
    pid = new_id()
    db.execute("""
        INSERT INTO meta_progresso (id, meta_id, user_id, quantidade, nota, created_at)
        VALUES (?,?,?,1,?,?)
    """, (pid, meta_id, g.current_user["id"], nota, now_iso()))
    audit("create", "meta_progresso", pid, {"meta_id": meta_id})
    db.commit()
    return jsonify({"id": pid}), 201


# ------------------------------------------------------------------
# Rotas — Relatórios Gerenciais
# Regra de visibilidade: vendedor só vê o próprio relatório (a API nem aceita
# vendedor_id/scope=all de quem não é admin). Admin vê o próprio por padrão,
# um vendedor específico via ?vendedor_id=, ou todo mundo via ?scope=all.
# ------------------------------------------------------------------
ORIGENS_VALIDAS = ("Outbound", "LinkedIn", "Indicações", "Eventos", "Inbound", "Outro")

MOTIVO_PERDA_LABEL = {
    "preco": "Preço", "concorrente": "Concorrente", "falta_funcionalidade": "Falta de funcionalidade/serviço",
    "sumiu_no_show": "Sumiu/não respondeu (no-show)", "sem_orcamento": "Sem orçamento",
    "timing_errado": "Timing errado", "outro": "Outro",
}

ETAPA_PROBABILIDADE = {"novo_lead": 0.15, "qualificacao": 0.35, "proposta_enviada": 0.55, "negociacao": 0.75}


@app.get("/api/reports/pipeline")
@login_required
def report_pipeline():
    db = get_db()
    campo_data = request.args.get("campo_data")  # "criacao" | "previsao" | None (foto atual)
    clause, params = report_scope_clause("user_id")

    filtro_data = ""
    if campo_data in ("criacao", "previsao"):
        try:
            inicio, fim = parse_period_from_request()
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
        if inicio and fim:
            col = "created_at" if campo_data == "criacao" else "data_prevista_fechamento"
            filtro_data = f" AND {col} BETWEEN ? AND ?"
            params = params + [inicio, fim + " 23:59:59"]

    rows = db.execute(f"""
        SELECT etapa_funil, COUNT(*) as quantidade, COALESCE(SUM(valor_estimado),0) as valor_total
        FROM deals WHERE status = 'aberto' AND categoria = 'padrao' {clause}{filtro_data}
        GROUP BY etapa_funil
    """, params).fetchall()
    por_etapa = {r["etapa_funil"]: dict(r) for r in rows}
    resultado = []
    for etapa in ETAPAS_ABERTAS:
        r = por_etapa.get(etapa, {"quantidade": 0, "valor_total": 0})
        qtd = r["quantidade"]
        resultado.append({
            "etapa": etapa, "etapa_label": ETAPA_LABEL_BACKEND[etapa],
            "quantidade": qtd, "valor_total": r["valor_total"],
            "ticket_medio": (r["valor_total"] / qtd) if qtd else 0,
        })
    return jsonify({"etapas": resultado})


@app.get("/api/reports/desempenho-equipe")
@login_required
def report_desempenho_equipe():
    db = get_db()
    try:
        inicio, fim = parse_period_from_request()
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    if not inicio:
        return jsonify({"error": "Informe o parâmetro 'periodo' para este relatório."}), 400
    clause, params = report_scope_clause("user_id")

    rows = db.execute(f"""
        SELECT deals.*, u.nome as vendedor_nome FROM deals
        LEFT JOIN users u ON u.id = deals.user_id
        WHERE status IN ('ganho','perdido') AND categoria = 'padrao' AND etapa_atualizada_em BETWEEN ? AND ? {clause}
    """, [inicio, fim + " 23:59:59"] + params).fetchall()

    por_vendedor = {}
    for d in rows:
        vid = d["user_id"]
        if vid not in por_vendedor:
            por_vendedor[vid] = {
                "vendedor_id": vid, "vendedor_nome": d["vendedor_nome"],
                "ganhos": 0, "perdidos": 0, "valor_ganho_total": 0.0, "valor_perdido_total": 0.0,
                "_dias_ciclo": [],
            }
        v = por_vendedor[vid]
        criado = datetime.strptime(d["created_at"], "%Y-%m-%d %H:%M:%S")
        fechado = datetime.strptime(d["etapa_atualizada_em"], "%Y-%m-%d %H:%M:%S")
        if d["status"] == "ganho":
            v["ganhos"] += 1
            v["valor_ganho_total"] += d["valor_estimado"] or 0
            v["_dias_ciclo"].append((fechado - criado).total_seconds() / 86400)
        else:
            v["perdidos"] += 1
            v["valor_perdido_total"] += d["valor_estimado"] or 0

    resultado = []
    for v in por_vendedor.values():
        total = v["ganhos"] + v["perdidos"]
        dias = v.pop("_dias_ciclo")
        resultado.append({
            **v,
            "taxa_conversao": round(v["ganhos"] / total * 100, 1) if total else 0,
            "ciclo_medio_dias": round(sum(dias) / len(dias), 1) if dias else None,
            "ticket_medio_ganho": round(v["valor_ganho_total"] / v["ganhos"], 2) if v["ganhos"] else 0,
        })
    resultado.sort(key=lambda r: r["valor_ganho_total"], reverse=True)
    return jsonify({"periodo": {"inicio": inicio, "fim": fim}, "vendedores": resultado})


@app.get("/api/reports/motivos-perda")
@login_required
def report_motivos_perda():
    db = get_db()
    try:
        inicio, fim = parse_period_from_request()
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    if not inicio:
        return jsonify({"error": "Informe o parâmetro 'periodo' para este relatório."}), 400
    clause, params = report_scope_clause("user_id")

    rows = db.execute(f"""
        SELECT motivo_perda, COUNT(*) as quantidade, COALESCE(SUM(valor_estimado),0) as valor_total
        FROM deals
        WHERE status = 'perdido' AND categoria = 'padrao' AND etapa_atualizada_em BETWEEN ? AND ? {clause}
        GROUP BY motivo_perda
        ORDER BY quantidade DESC
    """, [inicio, fim + " 23:59:59"] + params).fetchall()

    total_perdidos = sum(r["quantidade"] for r in rows)
    resultado = [{
        "motivo": r["motivo_perda"] or "nao_informado",
        "motivo_label": MOTIVO_PERDA_LABEL.get(r["motivo_perda"], "Não informado"),
        "quantidade": r["quantidade"], "valor_total": r["valor_total"],
        "percentual": round(r["quantidade"] / total_perdidos * 100, 1) if total_perdidos else 0,
    } for r in rows]
    return jsonify({"periodo": {"inicio": inicio, "fim": fim}, "total_perdidos": total_perdidos, "motivos": resultado})


@app.get("/api/reports/detalhes")
@login_required
def report_detalhes():
    """Drill-down dos relatórios: a lista de negócios por trás dos números.
    tipo=ganhos|perdidos (dentro do período) ou abertos (retrato de agora,
    ordenado do mais parado para o mais recente, com idade e dias na etapa).
    RBAC: vendedor só vê os próprios; admin usa ?scope=all ou ?vendedor_id=."""
    db = get_db()
    tipo = request.args.get("tipo")
    if tipo not in ("ganhos", "perdidos", "abertos"):
        return jsonify({"error": "Informe tipo=ganhos, perdidos ou abertos."}), 400
    clause, params = report_scope_clause("d.user_id")
    base = """
        SELECT d.id, d.titulo, d.valor_estimado, d.etapa_funil, d.status, d.categoria,
               d.motivo_perda, d.motivo_perda_detalhe, d.created_at, d.etapa_atualizada_em,
               d.produto_qtd, c.nome as cliente_nome, u.nome as vendedor_nome,
               pr.nome as produto_nome
        FROM deals d
        JOIN customers c ON c.id = d.customer_id
        LEFT JOIN users u ON u.id = d.user_id
        LEFT JOIN produtos pr ON pr.id = d.produto_id
    """
    if tipo == "abertos":
        rows = db.execute(base + f" WHERE d.status = 'aberto' {clause} ORDER BY d.etapa_atualizada_em ASC",
                          params).fetchall()
    else:
        try:
            inicio, fim = parse_period_from_request()
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
        if not inicio:
            return jsonify({"error": "Informe o parâmetro 'periodo' para este relatório."}), 400
        status = "ganho" if tipo == "ganhos" else "perdido"
        rows = db.execute(base + f" WHERE d.status = ? AND d.etapa_atualizada_em BETWEEN ? AND ? {clause}"
                                 " ORDER BY d.etapa_atualizada_em DESC",
                          [status, inicio, fim + " 23:59:59"] + params).fetchall()

    agora = datetime.now(timezone.utc)

    def _dias(ts):
        if not ts:
            return None
        try:
            dt = datetime.fromisoformat(str(ts).replace("Z", "")).replace(tzinfo=timezone.utc)
            return max(0, (agora - dt).days)
        except Exception:
            return None

    saida = []
    for r in rows:
        item = dict(r)
        item["motivo_label"] = (MOTIVO_PERDA_LABEL.get(r["motivo_perda"], "Não informado")
                                if r["status"] == "perdido" else None)
        item["dias_na_etapa"] = _dias(r["etapa_atualizada_em"])
        item["dias_vida"] = _dias(r["created_at"])
        saida.append(item)
    return jsonify({"negocios": saida, "admin": g.current_user["role"] == "admin"})


@app.get("/api/reports/atividades")
@login_required
def report_atividades():
    db = get_db()
    try:
        inicio, fim = parse_period_from_request()
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    if not inicio:
        return jsonify({"error": "Informe o parâmetro 'periodo' para este relatório."}), 400
    clause_tasks, params_tasks = report_scope_clause("user_id")
    clause_hist, params_hist = report_scope_clause("user_id")

    reunioes_agendadas = db.execute(f"""
        SELECT user_id, COUNT(*) c FROM tasks
        WHERE tipo_atividade = 'reuniao' AND created_at BETWEEN ? AND ? {clause_tasks}
        GROUP BY user_id
    """, [inicio, fim + " 23:59:59"] + params_tasks).fetchall()

    reunioes_realizadas = db.execute(f"""
        SELECT user_id, COUNT(*) c FROM tasks
        WHERE tipo_atividade = 'reuniao' AND executado = 1 AND executado_em BETWEEN ? AND ? {clause_tasks}
        GROUP BY user_id
    """, [inicio, fim + " 23:59:59"] + params_tasks).fetchall()

    follow_ups = db.execute(f"""
        SELECT user_id, COUNT(*) c FROM tasks
        WHERE tipo_atividade = 'follow_up' AND executado = 1 AND executado_em BETWEEN ? AND ? {clause_tasks}
        GROUP BY user_id
    """, [inicio, fim + " 23:59:59"] + params_tasks).fetchall()

    propostas = db.execute(f"""
        SELECT user_id, COUNT(*) c FROM deal_stage_history
        WHERE etapa_nova = 'proposta_enviada' AND data_transicao BETWEEN ? AND ? {clause_hist}
        GROUP BY user_id
    """, [inicio, fim + " 23:59:59"] + params_hist).fetchall()

    usuarios = {u["id"]: u["nome"] for u in db.execute("SELECT id, nome FROM users").fetchall()}
    por_vendedor = {}

    def _acumular(rows, campo):
        for r in rows:
            vid = r["user_id"]
            if vid not in por_vendedor:
                por_vendedor[vid] = {
                    "vendedor_id": vid, "vendedor_nome": usuarios.get(vid, "—"),
                    "reunioes_agendadas": 0, "reunioes_realizadas": 0,
                    "propostas_enviadas": 0, "follow_ups_realizados": 0,
                }
            por_vendedor[vid][campo] = r["c"]

    _acumular(reunioes_agendadas, "reunioes_agendadas")
    _acumular(reunioes_realizadas, "reunioes_realizadas")
    _acumular(propostas, "propostas_enviadas")
    _acumular(follow_ups, "follow_ups_realizados")

    resultado = sorted(por_vendedor.values(), key=lambda r: r["vendedor_nome"] or "")
    return jsonify({"periodo": {"inicio": inicio, "fim": fim}, "vendedores": resultado})


@app.get("/api/reports/forecast")
@login_required
def report_forecast():
    db = get_db()
    try:
        inicio, fim = parse_period_from_request()
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    if not inicio:
        return jsonify({"error": "Informe o parâmetro 'periodo' — o forecast é filtrado pela data prevista de fechamento."}), 400
    clause, params = report_scope_clause("user_id")

    rows = db.execute(f"""
        SELECT deals.*, c.nome as cliente_nome, u.nome as vendedor_nome FROM deals
        JOIN customers c ON c.id = deals.customer_id
        LEFT JOIN users u ON u.id = deals.user_id
        WHERE status = 'aberto' AND categoria = 'padrao' AND data_prevista_fechamento BETWEEN ? AND ? {clause}
        ORDER BY data_prevista_fechamento
    """, [inicio, fim] + params).fetchall()

    negocios = []
    valor_bruto_total = 0.0
    valor_ponderado_total = 0.0
    for d in rows:
        prob = ETAPA_PROBABILIDADE.get(d["etapa_funil"], 0)
        ponderado = (d["valor_estimado"] or 0) * prob
        valor_bruto_total += d["valor_estimado"] or 0
        valor_ponderado_total += ponderado
        negocios.append({
            "deal_id": d["id"], "titulo": d["titulo"], "cliente_nome": d["cliente_nome"],
            "vendedor_nome": d["vendedor_nome"], "etapa_funil": d["etapa_funil"],
            "valor_estimado": d["valor_estimado"], "probabilidade": prob, "valor_ponderado": round(ponderado, 2),
            "data_prevista_fechamento": d["data_prevista_fechamento"],
        })
    return jsonify({
        "periodo": {"inicio": inicio, "fim": fim},
        "valor_bruto_total": valor_bruto_total, "valor_ponderado_total": round(valor_ponderado_total, 2),
        "quantidade": len(negocios), "negocios": negocios,
    })


PRODUTOS_SOFTWARE = {
    "revele_momentos": "Revele Momentos",
    "revele_momentos_frontier": "Revele Momentos Frontier",
}
META_SOFTWARE = 1  # vendas ganhas esperadas por usuário no período (mensal)


@app.get("/api/reports/software")
@login_required
def report_software():
    """Relatório do produto estratégico 🎞 (software Revele Momentos):
    meta por usuário, totais e a lista completa das negociações de software.
    Visibilidade: vendedor só vê o próprio; admin vê todos, inclusive outros
    admins — a mesma regra dos demais relatórios."""
    db = get_db()
    try:
        inicio, fim = parse_period_from_request()
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    if not inicio:
        return jsonify({"error": "Informe o parâmetro 'periodo' para este relatório."}), 400
    clause, params = report_scope_clause("deals.user_id")

    rows = db.execute(f"""
        SELECT deals.*, c.nome as cliente_nome, u.nome as vendedor_nome
        FROM deals
        JOIN customers c ON c.id = deals.customer_id
        LEFT JOIN users u ON u.id = deals.user_id
        WHERE deals.categoria = 'software' AND deals.created_at BETWEEN ? AND ? {clause}
        ORDER BY deals.created_at DESC
    """, [inicio, fim + " 23:59:59"] + params).fetchall()

    # Painel de meta: no modo "todos", inclui TODOS os usuários ativos
    # (vendedores E admins — a meta vale para todo mundo); quem não ofereceu
    # nada aparece com meta ✗, que é justamente quem precisa agir.
    if g.current_user["role"] == "admin" and request.args.get("scope") == "all":
        usuarios_alvo = db.execute("SELECT id, nome, role FROM users WHERE ativo = 1 ORDER BY nome").fetchall()
    elif g.current_user["role"] == "admin" and request.args.get("vendedor_id"):
        usuarios_alvo = db.execute("SELECT id, nome, role FROM users WHERE id = ?",
                                   (request.args.get("vendedor_id"),)).fetchall()
    else:
        usuarios_alvo = db.execute("SELECT id, nome, role FROM users WHERE id = ?",
                                   (g.current_user["id"],)).fetchall()

    metas = {u["id"]: {"user_id": u["id"], "nome": u["nome"], "role": u["role"],
                       "ofertas": 0, "ganhas": 0, "perdidas": 0, "abertas": 0, "valor_ganho": 0.0}
             for u in usuarios_alvo}
    totais = {"ofertas": 0, "ganhas": 0, "perdidas": 0, "abertas": 0, "valor_ganho": 0.0}
    negocios = []
    for d in rows:
        m = metas.get(d["user_id"])
        chave = "ganhas" if d["status"] == "ganho" else ("perdidas" if d["status"] == "perdido" else "abertas")
        for alvo in ([m] if m is not None else []) + [totais]:
            alvo["ofertas"] += 1
            alvo[chave] += 1
            if d["status"] == "ganho":
                alvo["valor_ganho"] += d["valor_estimado"] or 0
        negocios.append({
            "deal_id": d["id"], "titulo": d["titulo"],
            "produto": PRODUTOS_SOFTWARE.get(d["produto_software"], d["produto_software"] or "—"),
            "cliente_nome": d["cliente_nome"], "vendedor_nome": d["vendedor_nome"],
            "valor_estimado": d["valor_estimado"], "status": d["status"],
            "etapa_funil": d["etapa_funil"], "created_at": d["created_at"],
        })
    usuarios = [{**m, "meta_ok": m["ganhas"] >= META_SOFTWARE} for m in metas.values()]

    # 🖨 Cobertura de oferta: clientes com impressora COMPATÍVEL que nunca
    # receberam uma oferta de software — a lista de abordagem da equipe.
    clause_cli, params_cli = report_scope_clause("c.responsavel_id")
    likes, params_likes = _clientes_sem_oferta_sql("c")
    cobertura_rows = db.execute(f"""
        SELECT c.id, c.nome, c.equipamentos, u.nome as responsavel_nome
        FROM customers c LEFT JOIN users u ON u.id = c.responsavel_id
        WHERE c.ativo = 1 AND ({likes}) {clause_cli}
          AND NOT EXISTS (SELECT 1 FROM deals d WHERE d.customer_id = c.id AND d.categoria = 'software')
        ORDER BY c.nome
    """, params_likes + params_cli).fetchall()
    cobertura = []
    for r in cobertura_rows:
        try:
            slugs = json.loads(r["equipamentos"] or "[]")
        except Exception:
            slugs = []
        cobertura.append({
            "customer_id": r["id"], "nome": r["nome"],
            "equipamentos": [EQUIPAMENTOS[s] for s in slugs if s in EQUIPAMENTOS],
            "responsavel_nome": r["responsavel_nome"],
        })

    return jsonify({"periodo": {"inicio": inicio, "fim": fim}, "meta_por_usuario": META_SOFTWARE,
                    "usuarios": usuarios, "totais": totais, "negocios": negocios,
                    "cobertura": cobertura})


@app.get("/api/reports/origem-leads")
@login_required
def report_origem_leads():
    db = get_db()
    try:
        inicio, fim = parse_period_from_request()
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    if not inicio:
        return jsonify({"error": "Informe o parâmetro 'periodo' para este relatório."}), 400
    clause, params = report_scope_clause("responsavel_id")

    clientes = db.execute(f"""
        SELECT id, origem FROM customers
        WHERE created_at BETWEEN ? AND ? {clause}
    """, [inicio, fim + " 23:59:59"] + params).fetchall()

    por_origem = {}
    for c in clientes:
        origem = c["origem"] or "Não informado"
        por_origem.setdefault(origem, {"origem": origem, "quantidade_leads": 0, "customer_ids": []})
        por_origem[origem]["quantidade_leads"] += 1
        por_origem[origem]["customer_ids"].append(c["id"])

    for origem, dados in por_origem.items():
        ids = dados.pop("customer_ids")
        if not ids:
            dados["quantidade_ganhos"] = 0
            dados["valor_total_ganho"] = 0
            dados["ticket_medio"] = 0
            continue
        placeholders = ",".join("?" for _ in ids)
        ganhos = db.execute(f"""
            SELECT COUNT(*) qtd, COALESCE(SUM(valor_estimado),0) valor
            FROM deals WHERE status = 'ganho' AND customer_id IN ({placeholders})
        """, ids).fetchone()
        dados["quantidade_ganhos"] = ganhos["qtd"]
        dados["valor_total_ganho"] = ganhos["valor"]
        dados["ticket_medio"] = round(ganhos["valor"] / ganhos["qtd"], 2) if ganhos["qtd"] else 0

    resultado = sorted(por_origem.values(), key=lambda r: r["valor_total_ganho"], reverse=True)
    return jsonify({"periodo": {"inicio": inicio, "fim": fim}, "origens": resultado})


# ------------------------------------------------------------------
# Rotas — Canal de notificação WhatsApp (apenas número de saída, sem leitura)
# ------------------------------------------------------------------
@app.get("/api/whatsapp/channels")
@login_required
def list_whatsapp_channels():
    db = get_db()
    rows = db.execute(
        "SELECT * FROM whatsapp_notification_channels WHERE user_id = ? ORDER BY created_at DESC",
        (g.current_user["id"],)
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/whatsapp/channels")
@login_required
def create_whatsapp_channel():
    body = request.get_json(force=True, silent=True) or {}
    numero = (body.get("numero_whatsapp") or "").strip()
    if not numero:
        return jsonify({"error": "Informe o número de WhatsApp."}), 400
    db = get_db()
    cid = new_id()
    db.execute("""
        INSERT INTO whatsapp_notification_channels (id, user_id, numero_whatsapp, provedor)
        VALUES (?,?,?,?)
    """, (cid, g.current_user["id"], numero, body.get("provedor", "whatsapp_cloud_api")))
    audit("create", "whatsapp_notification_channels", cid, body)
    db.commit()
    return jsonify({"id": cid}), 201


@app.delete("/api/whatsapp/channels/<channel_id>")
@login_required
def delete_whatsapp_channel(channel_id):
    db = get_db()
    db.execute(
        "DELETE FROM whatsapp_notification_channels WHERE id = ? AND user_id = ?",
        (channel_id, g.current_user["id"])
    )
    audit("delete", "whatsapp_notification_channels", channel_id)
    db.commit()
    return jsonify({"ok": True})


# ------------------------------------------------------------------
# Frontend estático
# ------------------------------------------------------------------
@app.get("/")
def index():
    return send_from_directory(app.static_folder, "dashboard.html")


# ------------------------------------------------------------------
# Bootstrap: cria schema + seed se o banco não existir
# ------------------------------------------------------------------
def migrate_missing_columns(conn):
    """Adiciona colunas novas em bancos criados por versões anteriores do app,
    sem apagar nenhum dado já cadastrado."""
    tabelas_e_colunas = {
        "customers": {
            "cpf_cnpj": "TEXT", "endereco": "TEXT", "cep": "TEXT", "cidade": "TEXT", "estado": "TEXT",
            "ativo": "INTEGER NOT NULL DEFAULT 1",
        },
        "delegated_tasks": {
            "grupo_id": "TEXT",
        },
        "deals": {
            "data_prevista_fechamento": "TEXT", "motivo_perda": "TEXT", "motivo_perda_detalhe": "TEXT",
        },
        "tasks": {
            "tipo_atividade": "TEXT NOT NULL DEFAULT 'outro'",
            "nota_conclusao": "TEXT",
        },
        "deal_notes": {
            "tipo": "TEXT NOT NULL DEFAULT 'nota'",
            "tem_anexo": "INTEGER NOT NULL DEFAULT 0",
        },
        "deal_stage_history": {
            "motivo_perda": "TEXT",
            "motivo_perda_detalhe": "TEXT",
        },
        "customers": {
            "recompra_dias": "INTEGER",
            "proxima_recompra": "TEXT",
            "equipamentos": "TEXT",
            "rolos_mes_media": "INTEGER",
        },
        "deals": {
            "origem_recompra": "INTEGER NOT NULL DEFAULT 0",
            "categoria": "TEXT NOT NULL DEFAULT 'padrao'",
            "produto_software": "TEXT",
            "produto_id": "TEXT",
            "produto_qtd": "INTEGER",
        },
        "metas": {
            "alvo_produto_id": "TEXT",
            "editada_por": "TEXT",
            "editada_em": "TEXT",
            "escopo_users": "TEXT",
        },
        "condicoes_pagamento": {
            "acrescimo_pct": "REAL NOT NULL DEFAULT 0",
            "simples": "INTEGER NOT NULL DEFAULT 0",
        },
        "deals": {
            "condicao_pagamento_id": "TEXT",
        },
        "deal_itens": {
            "aprovado": "INTEGER NOT NULL DEFAULT 1",
            "liberacao_id": "TEXT",
        },
        "produtos": {
            "seq": "INTEGER",
            "categoria": "TEXT",
            "equipamento": "TEXT",
            "embalagem": "TEXT",
            "preco_tabela": "REAL",
            "preco_limite": "REAL",
            "desconto_max": "REAL",
            "status_comercial": "TEXT",
            "ofertavel": "INTEGER NOT NULL DEFAULT 1",
        },
    }
    for tabela, colunas in tabelas_e_colunas.items():
        colunas_existentes = {row["name"] for row in conn.execute(f"PRAGMA table_info({tabela})").fetchall()}
        for coluna, tipo in colunas.items():
            if coluna not in colunas_existentes:
                conn.execute(f"ALTER TABLE {tabela} ADD COLUMN {coluna} {tipo}")

    # Migração especial: remover o UNIQUE de customers.whatsapp_id em bancos
    # antigos (a mesma pessoa pode ter várias empresas com o mesmo número).
    # SQLite não solta constraint por ALTER: é preciso reconstruir a tabela.
    precisa_rebuild = False
    for idx in conn.execute("PRAGMA index_list(customers)").fetchall():
        if idx["unique"]:
            cols = [r["name"] for r in conn.execute(f"PRAGMA index_info({idx['name']})").fetchall()]
            if cols == ["whatsapp_id"]:
                precisa_rebuild = True
    if precisa_rebuild:
        conn.execute("PRAGMA foreign_keys = OFF")
        conn.execute("""
            CREATE TABLE customers_nova (
                id                  TEXT PRIMARY KEY,
                nome                TEXT NOT NULL,
                whatsapp_id         TEXT,
                telefone            TEXT,
                email               TEXT,
                cpf_cnpj            TEXT,
                endereco            TEXT,
                cep                 TEXT,
                cidade              TEXT,
                estado              TEXT,
                ativo               INTEGER NOT NULL DEFAULT 1,
                data_ultima_compra  TEXT,
                status_fidelidade   TEXT NOT NULL CHECK(status_fidelidade IN ('novo','recorrente','vip','inativo','perdido')) DEFAULT 'novo',
                recompra_dias       INTEGER,
                proxima_recompra    TEXT,
                equipamentos        TEXT,
                rolos_mes_media     INTEGER,
                responsavel_id      TEXT REFERENCES users(id) ON DELETE SET NULL,
                origem              TEXT,
                observacoes         TEXT,
                created_at          TEXT NOT NULL DEFAULT (datetime('now')),
                updated_at          TEXT NOT NULL DEFAULT (datetime('now'))
            )
        """)
        cols_novas = [r["name"] for r in conn.execute("PRAGMA table_info(customers_nova)").fetchall()]
        cols_velhas = [r["name"] for r in conn.execute("PRAGMA table_info(customers)").fetchall()]
        comuns = ", ".join(col for col in cols_novas if col in cols_velhas)
        conn.execute(f"INSERT INTO customers_nova ({comuns}) SELECT {comuns} FROM customers")
        conn.execute("DROP TABLE customers")
        conn.execute("ALTER TABLE customers_nova RENAME TO customers")
        conn.execute("PRAGMA foreign_keys = ON")
        conn.commit()

    # 💳 As 3 condições SIMPLIFICADAS do orçamento (ids fixos: sobrevivem a
    # reimportações e mantêm os orçamentos antigos apontando certo).
    # A tabela detalhada importada da planilha permanece como consulta.
    for cid, forma, condicao, regra, pct, ordem in (
        ("cond-pix", "PIX/TED", "À vista", "", 0.0, 1),
        ("cond-cartao", "Cartão de crédito", "Até 10x, com acréscimo de 2%", "", 2.0, 2),
        ("cond-boleto", "Boleto", "Prazos conforme o valor do pedido",
         "Sujeito à aprovação do financeiro — consulte as faixas na tabela 💳 do catálogo", 0.0, 3),
    ):
        conn.execute("""
            INSERT OR REPLACE INTO condicoes_pagamento
                (id, perfil, forma, condicao, regra, eh_nota, ordem, acrescimo_pct, simples, atualizado_em)
            VALUES (?,?,?,?,?,0,?,?,1,datetime('now'))
        """, (cid, "Todos", forma, condicao, regra, ordem, pct))
    conn.commit()


def init_db_if_needed():
    fresh = False
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    with open("schema.sql", encoding="utf-8") as f:
        conn.executescript(f.read())
    migrate_missing_columns(conn)
    already_seeded = conn.execute("SELECT COUNT(*) c FROM users").fetchone()["c"] > 0
    if not already_seeded:
        fresh = True
        if os.environ.get("CRM_SEED_DEMO", "1") == "0":
            # PRODUÇÃO (CRM_SEED_DEMO=0): nada de dados de exemplo — cria só o
            # administrador inicial. Troque a senha no primeiro acesso (Equipe).
            conn.execute(
                "INSERT INTO users (id, nome, email, senha_hash, role) VALUES (?,?,?,?,?)",
                (new_id(),
                 os.environ.get("CRM_ADMIN_NOME", "Administrador"),
                 os.environ.get("CRM_ADMIN_EMAIL", "admin@digimagem.com"),
                 generate_password_hash(os.environ.get("CRM_ADMIN_SENHA", "trocar123")),
                 "admin"))
        else:
            seed(conn)
    # 🎯 Bootstrap único: se NUNCA existiu meta alguma (nem excluída), cria a
    # meta padrão do software — a antiga meta fixa do código, agora editável
    # e excluível pelo administrador na tela Metas.
    tem_meta = conn.execute("SELECT COUNT(*) c FROM metas").fetchone()["c"] > 0
    if not tem_meta:
        admin = conn.execute("SELECT id FROM users WHERE role = 'admin' ORDER BY created_at LIMIT 1").fetchone()
        if admin:
            conn.execute("""
                INSERT INTO metas (id, titulo, descricao, tipo, alvo_vendas, quantidade,
                                   apuracao, escopo, periodo_tipo, criado_por)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (new_id(), "Vender 1 assinatura do software Revele Momentos",
                  "Oferecer e vender o Revele Momentos ou o Revele Momentos Frontier aos clientes de impressoras ASK/Frontier.",
                  "vendas", "software_qualquer", 1, "individual", "todos", "mensal", admin["id"]))
    conn.commit()
    conn.close()
    return fresh


def seed(conn):
    def uid():
        return uuid.uuid4().hex

    admin1, admin2 = uid(), uid()
    carlos_vendas, ana, tiago = uid(), uid(), uid()

    conn.execute("INSERT INTO users (id, nome, email, senha_hash, role) VALUES (?,?,?,?,?)",
                 (admin1, "Admin Master 1", "admin1@digimagem.com", generate_password_hash("admin123"), "admin"))
    conn.execute("INSERT INTO users (id, nome, email, senha_hash, role) VALUES (?,?,?,?,?)",
                 (admin2, "Admin Master 2", "admin2@digimagem.com", generate_password_hash("admin123"), "admin"))
    conn.execute("INSERT INTO users (id, nome, email, senha_hash, role) VALUES (?,?,?,?,?)",
                 (carlos_vendas, "Carlos", "carlos.vendas@lojadigimagem.com.br", generate_password_hash("vendas123"), "vendedor"))
    conn.execute("INSERT INTO users (id, nome, email, senha_hash, role) VALUES (?,?,?,?,?)",
                 (ana, "Ana", "ana@lojadigimagem.com.br", generate_password_hash("vendas123"), "vendedor"))
    conn.execute("INSERT INTO users (id, nome, email, senha_hash, role) VALUES (?,?,?,?,?)",
                 (tiago, "Tiago", "tiago@lojadigimagem.com.br", generate_password_hash("vendas123"), "vendedor"))

    def add_customer(nome, whatsapp, status, dias_sem_comprar, responsavel, origem="Indicações"):
        cid = uid()
        data_compra = (datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=dias_sem_comprar)).strftime("%Y-%m-%d") if dias_sem_comprar is not None else None
        conn.execute("""
            INSERT INTO customers (id, nome, whatsapp_id, data_ultima_compra, status_fidelidade, responsavel_id, origem)
            VALUES (?,?,?,?,?,?,?)
        """, (cid, nome, whatsapp, data_compra, status, responsavel, origem))
        return cid

    joao = add_customer("João Pereira", "5548999990001", "vip", 18, carlos_vendas, "Indicações")
    studio_nova = add_customer("Studio Nova", "5548999990002", "recorrente", 46, ana, "LinkedIn")
    bruno = add_customer("Bruno Lima", "5548999990003", "novo", None, carlos_vendas, "Inbound")
    fernanda = add_customer("Fernanda Dias", "5548999990004", "novo", None, ana, "Outbound")
    atelie = add_customer("Ateliê Prime", "5548999990005", "vip", 5, tiago, "Eventos")
    ricardo = add_customer("Ricardo Alves", "5548999990006", "novo", None, tiago, "LinkedIn")
    perdida1 = add_customer("Contatos Perdidos ME", "5548999990007", "perdido", None, carlos_vendas, "Outbound")
    ganha1 = add_customer("Foto Prime Estúdio", "5548999990008", "recorrente", None, tiago, "Indicações")

    def add_deal(customer_id, user_id, titulo, etapa, valor, horas_parado=0, status="aberto",
                 motivo_perda=None, dias_previsao=None, dias_atras_fechamento=None):
        did = uid()
        if dias_atras_fechamento is not None:
            etapa_ts = (datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=dias_atras_fechamento)).strftime("%Y-%m-%d %H:%M:%S")
            criado_ts = (datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=dias_atras_fechamento + 12)).strftime("%Y-%m-%d %H:%M:%S")
        else:
            etapa_ts = (datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(hours=horas_parado)).strftime("%Y-%m-%d %H:%M:%S")
            criado_ts = etapa_ts
        previsao = (datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(days=dias_previsao)).strftime("%Y-%m-%d") if dias_previsao is not None else None
        conn.execute("""
            INSERT INTO deals (id, customer_id, user_id, titulo, etapa_funil, valor_estimado, status,
                etapa_atualizada_em, created_at, data_prevista_fechamento, motivo_perda)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (did, customer_id, user_id, titulo, etapa, valor, status, etapa_ts, criado_ts, previsao, motivo_perda))
        conn.execute("""
            INSERT INTO deal_stage_history (id, deal_id, etapa_anterior, etapa_nova, user_id, data_transicao)
            VALUES (?,?,NULL,?,?,?)
        """, (uid(), did, etapa, user_id, criado_ts))
        if etapa != "novo_lead" and dias_atras_fechamento is None:
            # registra também a passagem por "proposta_enviada" pra alimentar o relatório de atividades
            conn.execute("""
                INSERT INTO deal_stage_history (id, deal_id, etapa_anterior, etapa_nova, user_id, data_transicao)
                VALUES (?,?,?,?,?,?)
            """, (uid(), did, "novo_lead", etapa, user_id, etapa_ts))
        return did

    deal_joao = add_deal(joao, carlos_vendas, "Ensaio corporativo — João Pereira", "qualificacao", 7400, horas_parado=52, dias_previsao=10)
    add_deal(bruno, carlos_vendas, "Pacote casamento — Bruno Lima", "proposta_enviada", 6200, horas_parado=61, dias_previsao=5)
    add_deal(studio_nova, ana, "Contrato mensal — Studio Nova", "negociacao", 12000, horas_parado=10, dias_previsao=15)
    add_deal(fernanda, ana, "Book individual — Fernanda Dias", "novo_lead", 3200, horas_parado=6)
    add_deal(atelie, tiago, "Catálogo produtos — Ateliê Prime", "negociacao", 18500, horas_parado=20, dias_previsao=8)
    add_deal(ricardo, tiago, "Ensaio família — Ricardo Alves", "novo_lead", 5900, horas_parado=2)
    # Negócios já fechados, pra alimentar os relatórios de Desempenho, Motivos de Perda e Origem de Leads
    add_deal(perdida1, carlos_vendas, "Cobertura evento — Contatos Perdidos ME", "fechado_perdido", 4500,
              status="perdido", motivo_perda="preco", dias_atras_fechamento=6)
    add_deal(ganha1, tiago, "Ensaio produtos — Foto Prime Estúdio", "fechado_ganho", 8900,
              status="ganho", dias_atras_fechamento=3)

    # task vencida (compromisso esquecido) ligada ao João Pereira
    conn.execute("""
        INSERT INTO tasks (id, deal_id, customer_id, user_id, descricao, tipo_atividade, data_lembrete, executado)
        VALUES (?,?,?,?,?,'follow_up',?,0)
    """, (uid(), deal_joao, joao, carlos_vendas, 'Retornar contato — "te ligo na terça"',
          (datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=2)).strftime("%Y-%m-%d %H:%M:%S")))

    # task futura, ainda não vencida
    conn.execute("""
        INSERT INTO tasks (id, deal_id, customer_id, user_id, descricao, tipo_atividade, data_lembrete, executado)
        VALUES (?,?,?,?,?,'proposta',?,0)
    """, (uid(), None, atelie, tiago, "Enviar proposta revisada",
          (datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(days=2)).strftime("%Y-%m-%d %H:%M:%S")))

    # reunião já realizada, pra alimentar o Relatório de Atividades Comerciais
    conn.execute("""
        INSERT INTO tasks (id, deal_id, customer_id, user_id, descricao, tipo_atividade, data_lembrete, executado, executado_em)
        VALUES (?,?,?,?,?,'reuniao',?,1,?)
    """, (uid(), deal_joao, joao, carlos_vendas, "Reunião de diagnóstico com o cliente",
          (datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=3)).strftime("%Y-%m-%d %H:%M:%S"),
          (datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=3)).strftime("%Y-%m-%d %H:%M:%S")))

    # follow-up já realizado
    conn.execute("""
        INSERT INTO tasks (id, deal_id, customer_id, user_id, descricao, tipo_atividade, data_lembrete, executado, executado_em)
        VALUES (?,?,?,?,?,'follow_up',?,1,?)
    """, (uid(), None, ganha1, tiago, "Follow-up pós-fechamento",
          (datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S"),
          (datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S")))

    conn.commit()


# Roda sempre que o módulo é carregado — tanto no `python app.py` local quanto
# quando o Gunicorn importa `app` em produção (Gunicorn nunca executa o bloco
# `if __name__ == "__main__":` abaixo, então a inicialização precisa estar aqui fora).
_fresh_db = init_db_if_needed()
if _fresh_db:
    print("Banco criado e populado com dados de exemplo (crm.db).")

if __name__ == "__main__":
    porta = int(os.environ.get("PORT", 5000))
    modo_debug = os.environ.get("FLASK_DEBUG") == "1"
    print(f"CRM-Digimagem rodando em http://localhost:{porta}")
    app.run(debug=modo_debug, host="0.0.0.0", port=porta, use_reloader=False)
